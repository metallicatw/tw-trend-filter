# -*- coding: utf-8 -*-
"""台股順勢交易篩選系統 V3.1 ─ 篩選引擎與報告產生器。

從本機單機版（`TW_Stock_Trend_Following_Trading_Filter_PC_V3.1.py`）移植而來。
篩選邏輯、指標公式、Excel 版面、互動線圖的外觀一個位元組都沒有改；改掉的是
「這支程式假設它跑在誰的桌面上」那部分：

* 開頭那段自動 `pip install` 拿掉了 —— 相依由 `requirements.txt` 宣告，CI 裝一次。
  一支排程在 runner 上偷偷升級套件，是那種第 87 天才會爆、而且爆的時候看不出
  是誰動的手的問題。
* 字型候選清單多了 Linux 的 Noto Sans CJK —— matplotlib 找不到中文字型不會報錯，
  只會把每一個中文字畫成豆腐塊，然後你在 Excel 裡才發現。
* 跑完不再 `os.startfile` / 開瀏覽器 / 等 `input()` —— runner 上沒有人按 Enter。
* 輸出路徑、股票池上限、個股頁連結、K 線保留年數都變成參數。

`run()` 是唯一的進入點，`__main__.py` 把命令列翻譯成它的參數。
"""

from __future__ import annotations

import warnings
warnings.filterwarnings('ignore')

import os, sys, datetime, platform, tempfile, subprocess as _sp
from io import StringIO

import requests
import yfinance as yf
from curl_cffi import requests as _cffi_requests
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.font_manager import FontProperties
from matplotlib.ticker import FuncFormatter
from matplotlib.patches import Rectangle

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.worksheet.table import Table, TableStyleInfo

# Yahoo Finance 會依 TLS 指紋辨識非瀏覽器連線並回傳 401 Invalid Crumb，
# 這裡用 curl_cffi 模擬 Chrome 連線來繞過此限制。
_YF_SESSION = _cffi_requests.Session(impersonate='chrome')

#: 這支程式版本號。出現在 Excel 抬頭、HTML 標題與 CLI 的 `--version`。
VERSION = 'V3.1'


# ===============================================================
# 1. 字型與樣式設定
# ===============================================================
# Windows 的候選維持原樣（連同那個雙反斜線的寫法——Windows API 會把重複的分隔符
# 正規化掉，所以它一直是能用的，改它只是製造一個和本機版不一樣的地方）；後面接
# 上 Linux 的候選，因為排程跑在 ubuntu runner 上。
#
# 字型找不到不會讓 matplotlib 報錯，只會把每一個中文字畫成一個空心方框——那種
# 錯誤要等到有人打開 Excel 看到滿頁豆腐才會發現，所以 workflow 裡會 apt 裝
# fonts-noto-cjk，而這裡列出它裝到哪。
CJK_FONT_CANDIDATES = [
    r'C:\\Windows\\Fonts\\msjh.ttc',
    r'C:\\Windows\\Fonts\\msjhbd.ttc',
    r'C:\\Windows\\Fonts\\mingliu.ttc',
    r'C:\\Windows\\Fonts\\kaiu.ttf',
    r'C:\Windows\Fonts\msjh.ttc',
    r'C:\Windows\Fonts\msjhbd.ttc',
    '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
    '/usr/share/fonts/opentype/noto/NotoSerifCJK-Regular.ttc',
    '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
    '/System/Library/Fonts/PingFang.ttc',
]
CJK_FONT_BOLD_CANDIDATES = [
    r'C:\\Windows\\Fonts\\msjhbd.ttc',
    r'C:\\Windows\\Fonts\\msjh.ttc',
    r'C:\\Windows\\Fonts\\mingliu.ttc',
    r'C:\Windows\Fonts\msjhbd.ttc',
    '/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc',
    '/usr/share/fonts/truetype/noto/NotoSansCJK-Bold.ttc',
    '/System/Library/Fonts/PingFang.ttc',
]

def safe_font_properties():
    for fp in CJK_FONT_CANDIDATES:
        if os.path.exists(fp):
            try: return FontProperties(fname=fp)
            except: pass
    return FontProperties()

def safe_font_properties_bold():
    for fp in CJK_FONT_BOLD_CANDIDATES:
        if os.path.exists(fp):
            try: return FontProperties(fname=fp)
            except: pass
    return FontProperties(weight='bold')

FP      = safe_font_properties()
FP_BOLD = safe_font_properties_bold()

matplotlib.rcParams['font.sans-serif'] = [
    'Noto Sans CJK TC','Noto Sans CJK JP','Noto Sans CJK SC',
    'Taipei Sans TC Beta','Microsoft JhengHei','SimHei',
    'Arial Unicode MS','DejaVu Sans'
]
matplotlib.rcParams['font.family']        = 'sans-serif'
matplotlib.rcParams['axes.unicode_minus'] = False

def mkfont(name='Microsoft JhengHei', size=11, bold=False, color='000000', italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)
def mkfill(hx):
    return PatternFill('solid', fgColor=hx)

THIN = Side(style='thin',   color='B0BEC5')
MED  = Side(style='medium', color='78909C')
def border_all(thin=True):
    s = THIN if thin else MED
    return Border(left=s, right=s, top=s, bottom=s)

ALN_CC  = Alignment(horizontal='center', vertical='center')
ALN_LC  = Alignment(horizontal='left',   vertical='center', indent=1)
ALN_RC  = Alignment(horizontal='right',  vertical='center')
ALN_LCW = Alignment(horizontal='left',   vertical='center', indent=1, wrap_text=True)
ALN_CCW = Alignment(horizontal='center', vertical='center', wrap_text=True)

C_NAVY   = '0D1B2A'
C_TEAL   = '0E6655'
C_HEADER = '1A3A5C'
C_ROW1   = 'EBF5FB'
C_ROW2   = 'FDFEFE'
C_ACCENT = 'E74C3C'
C_STOPBG = 'FDEDEC'

COLORS = {
    'bg':      '#0d1117',
    'axes_bg': '#161b22',
    'grid':    '#21262d',
    'text':    '#e6edf3',
    'ma20':    '#FFD700',
    'ma60':    '#00BFFF',
    'vol_up':  '#3fb950',
    'vol_dn':  '#f85149',
    'stop':    '#ff4444',
}

# ===============================================================
# 2. 讀取全市場股票清單（ISIN 主來源 → OpenAPI 備援）
# ===============================================================
def load_tw_stock_universe():
    all_rows    = []
    isin_industry = {}          # code → 中文產業別（直接從 ISIN HTML 第 4 欄取得）

    # A. TWSE ISIN HTML（主來源）
    for url, suffix in [
        ('https://isin.twse.com.tw/isin/C_public.jsp?strMode=2', '.TW'),
        ('https://isin.twse.com.tw/isin/C_public.jsp?strMode=4', '.TWO'),
    ]:
        try:
            resp = requests.get(url, timeout=20, headers={'User-Agent':'Mozilla/5.0'})
            resp.encoding = 'big5'
            tables = pd.read_html(StringIO(resp.text), header=0)
            df = tables[0].dropna(how='all').astype(str)

            # 第 0 欄：代號＋名稱（e.g. "3706  神達"）
            extracted = df.iloc[:, 0].str.extract(r'^(\d{4})\s+(.+)$')
            extracted.columns = ['code', 'name']
            mask = extracted['code'].notna()
            extracted = extracted[mask].copy()
            extracted['ticker'] = extracted['code'] + suffix
            all_rows.append(extracted)

            # 第 4 欄：產業別（中文，e.g. "電腦及週邊設備業"）
            if df.shape[1] > 4:
                ind_col = df.iloc[:, 4]
                for idx in extracted.index:
                    code = extracted.loc[idx, 'code']
                    raw  = str(ind_col.loc[idx]).strip() if idx in ind_col.index else ''
                    if raw and raw not in ('nan', 'NaN', 'None', '', '-', '產業別'):
                        isin_industry[code] = raw

            print(f'✅ ISIN {suffix}：{len(extracted)} 檔  (取得產業別 {sum(1 for k in extracted["code"] if k in isin_industry)} 筆)')
        except Exception as e:
            print(f'⚠️ ISIN {suffix} 失敗：{e}')

    if all_rows:
        all_df = pd.concat(all_rows, ignore_index=True).drop_duplicates('ticker')
        return all_df['ticker'].tolist(), all_df.set_index('code')['name'].to_dict(), isin_industry

    # B. TWSE OpenAPI（備援，無產業別）
    print('🔄 改用 OpenAPI 備援...')
    rows = []
    for url, sfx in [
        ('https://openapi.twse.com.tw/v1/opendata/t187ap03_L', '.TW'),
        ('https://openapi.twse.com.tw/v1/opendata/t187ap03_O', '.TWO'),
    ]:
        try:
            resp = requests.get(url, timeout=20)
            resp.raise_for_status()
            for r in resp.json():
                code = str(r.get('公司代號','')).strip()
                name = str(r.get('公司簡稱','')).strip()
                ind  = str(r.get('產業別', '')).strip()
                if code.isdigit() and len(code) == 4:
                    rows.append({'code': code, 'name': name, 'ticker': code + sfx})
                    if ind and ind not in ('nan', 'NaN', 'None', ''):
                        isin_industry[code] = ind
        except Exception as e:
            print(f'⚠️ OpenAPI {sfx} 失敗：{e}')
    all_df = pd.DataFrame(rows).drop_duplicates('ticker')
    print(f'✅ OpenAPI：{len(all_df)} 檔')
    return all_df['ticker'].tolist(), all_df.set_index('code')['name'].to_dict(), isin_industry



# ===============================================================
# 產業類別對照表（依 TWSE/TPEx 官方中文產業分類名稱）
# ===============================================================
# ================================================================
# 產業代號 → 中文名稱對照（TWSE/TPEx 官方分類代號）
# ================================================================
INDUSTRY_CODE_TO_NAME = {
    '01':'水泥工業',   '02':'食品工業',   '03':'塑膠工業',
    '04':'紡織纖維',   '05':'電機機械',   '06':'電器電纜',
    '07':'化學工業',   '08':'生技醫療業', '09':'玻璃陶瓷',
    '10':'造紙工業',   '11':'鋼鐵工業',   '12':'橡膠工業',
    '13':'汽車工業',   '14':'半導體業',   '15':'電腦及週邊設備業',
    '16':'光電業',     '17':'通訊網路業', '18':'電子零組件業',
    '19':'電子通路業', '20':'資訊服務業', '21':'其他電子業',
    '22':'建材營造業', '23':'航運業',     '24':'觀光餐旅業',
    '25':'金融保險業', '26':'貿易百貨業', '27':'油電燃氣業',
    '28':'電信服務業', '29':'綜合',       '30':'其他',
    '31':'存託憑證',   '32':'文化創意業', '33':'農業科技業',
    '34':'電子商務',   '35':'綠能環保',   '36':'數位雲端',
    '37':'運動休閒',   '38':'居家生活',
}

# ================================================================
# 產業類別靜態對照表（依股票代碼，每個代碼唯一）
# ================================================================
INDUSTRY_MAP = {
    '1101':'水泥工業','1102':'水泥工業','1103':'水泥工業','1104':'水泥工業',
    '1108':'水泥工業','1109':'水泥工業','1110':'水泥工業',
    '1201':'食品工業','1203':'食品工業','1210':'食品工業','1213':'食品工業',
    '1214':'食品工業','1215':'食品工業','1216':'食品工業','1217':'食品工業',
    '1218':'食品工業','1219':'食品工業','1220':'食品工業','1225':'食品工業',
    '1227':'食品工業','1229':'食品工業','1231':'食品工業','1232':'食品工業',
    '1233':'食品工業','1234':'食品工業','1235':'食品工業','1236':'食品工業',
    '1256':'食品工業',
    '1301':'塑膠工業','1303':'塑膠工業','1304':'塑膠工業','1305':'塑膠工業',
    '1307':'塑膠工業','1308':'塑膠工業','1309':'塑膠工業','1310':'塑膠工業',
    '1312':'塑膠工業','1313':'塑膠工業','1314':'塑膠工業','1315':'塑膠工業',
    '1316':'塑膠工業','1317':'塑膠工業','1318':'塑膠工業','1319':'塑膠工業',
    '1321':'塑膠工業','1323':'塑膠工業','1324':'塑膠工業','1325':'塑膠工業',
    '1326':'塑膠工業',
    '1402':'紡織纖維','1403':'紡織纖維','1404':'紡織纖維','1405':'紡織纖維',
    '1406':'紡織纖維','1414':'紡織纖維','1416':'紡織纖維','1417':'紡織纖維',
    '1418':'紡織纖維','1419':'紡織纖維','1423':'紡織纖維','1424':'紡織纖維',
    '1425':'紡織纖維','1426':'紡織纖維','1427':'紡織纖維','1429':'紡織纖維',
    '1431':'紡織纖維','1432':'紡織纖維','1433':'紡織纖維','1434':'紡織纖維',
    '1435':'紡織纖維','1436':'紡織纖維','1437':'紡織纖維','1438':'紡織纖維',
    '1439':'紡織纖維','1440':'紡織纖維','1441':'紡織纖維','1442':'紡織纖維',
    '1443':'紡織纖維','1444':'紡織纖維','1445':'紡織纖維','1446':'紡織纖維',
    '1447':'紡織纖維','1448':'紡織纖維','1449':'紡織纖維','1451':'紡織纖維',
    '1452':'紡織纖維','1453':'紡織纖維','1454':'紡織纖維','1455':'紡織纖維',
    '1456':'紡織纖維','1457':'紡織纖維','1459':'紡織纖維','1460':'紡織纖維',
    '1461':'紡織纖維','1462':'紡織纖維','1463':'紡織纖維','1464':'紡織纖維',
    '1465':'紡織纖維','1466':'紡織纖維','1467':'紡織纖維','1468':'紡織纖維',
    '1469':'紡織纖維','1470':'紡織纖維','1471':'紡織纖維','1472':'紡織纖維',
    '1473':'紡織纖維','1474':'紡織纖維','1475':'紡織纖維','1476':'紡織纖維',
    '1477':'紡織纖維','1478':'紡織纖維',
    '1503':'電機機械','1504':'電機機械','1506':'電機機械','1507':'電機機械',
    '1512':'電機機械','1513':'電機機械','1514':'電機機械','1515':'電機機械',
    '1516':'電機機械','1517':'電機機械','1519':'電機機械','1521':'電機機械',
    '1522':'電機機械','1523':'電機機械','1524':'電機機械','1525':'電機機械',
    '1526':'電機機械','1527':'電機機械','1528':'電機機械','1529':'電機機械',
    '1530':'電機機械','1531':'電機機械','1532':'電機機械','1533':'電機機械',
    '1535':'電機機械','1536':'電機機械','1537':'電機機械','1538':'電機機械',
    '1539':'電機機械','1540':'電機機械','1541':'電機機械','1543':'電機機械',
    '1544':'電機機械','1545':'電機機械','1546':'電機機械','1547':'電機機械',
    '1548':'電機機械','1549':'電機機械','1550':'電機機械','1551':'電機機械',
    '1552':'電機機械','1553':'電機機械','1554':'電機機械','1555':'電機機械',
    '1556':'電機機械','1558':'電機機械','1560':'電機機械','1561':'電機機械',
    '1562':'電機機械','1563':'電機機械','1564':'電機機械','1565':'電機機械',
    '1566':'電機機械','1568':'電機機械','1569':'電機機械','1570':'電機機械',
    '1571':'電機機械','1572':'電機機械','1573':'電機機械','1574':'電機機械',
    '1575':'電機機械','1576':'電機機械','1577':'電機機械','1579':'電機機械',
    '1580':'電機機械','1582':'電機機械','1583':'電機機械','1584':'電機機械',
    '1585':'電機機械','1586':'電機機械','1587':'電機機械','1589':'電機機械',
    '1590':'電機機械','1591':'電機機械','1592':'電機機械','1593':'電機機械',
    '1597':'電機機械','1598':'電機機械','1599':'電機機械',
    '1603':'電器電纜','1604':'電器電纜','1605':'電器電纜','1608':'電器電纜',
    '1609':'電器電纜','1611':'電器電纜','1612':'電器電纜','1613':'電器電纜',
    '1614':'電器電纜','1615':'電器電纜','1616':'電器電纜','1617':'電器電纜',
    '1618':'電器電纜','1619':'電器電纜','1620':'電器電纜',
    '1701':'化學工業','1702':'化學工業','1704':'化學工業','1705':'化學工業',
    '1706':'化學工業','1707':'化學工業','1708':'化學工業','1709':'化學工業',
    '1710':'化學工業','1711':'化學工業','1712':'化學工業','1713':'化學工業',
    '1714':'化學工業','1715':'化學工業','1717':'化學工業','1718':'化學工業',
    '1719':'化學工業','1720':'化學工業','1721':'化學工業','1722':'化學工業',
    '1723':'化學工業','1724':'化學工業','1725':'化學工業','1726':'化學工業',
    '1727':'化學工業','1730':'化學工業','1731':'化學工業','1732':'化學工業',
    '1733':'化學工業','1734':'化學工業','1735':'化學工業','1736':'化學工業',
    '1737':'化學工業','1738':'化學工業','1741':'化學工業','1742':'化學工業',
    '1752':'化學工業','1753':'化學工業','1754':'化學工業','1755':'化學工業',
    '1760':'化學工業','1762':'化學工業','1773':'化學工業',
    '1786':'生技醫療業','1788':'生技醫療業','1789':'生技醫療業',
    '1795':'生技醫療業','1796':'生技醫療業','1802':'玻璃陶瓷',
    '1805':'生技醫療業','1806':'玻璃陶瓷','1807':'玻璃陶瓷',
    '1808':'玻璃陶瓷','1809':'玻璃陶瓷',
    '1901':'造紙工業','1903':'造紙工業','1904':'造紙工業','1905':'造紙工業',
    '1906':'造紙工業','1907':'造紙工業',
    '2002':'鋼鐵工業','2006':'鋼鐵工業','2007':'鋼鐵工業','2008':'鋼鐵工業',
    '2009':'鋼鐵工業','2010':'鋼鐵工業','2011':'鋼鐵工業','2012':'鋼鐵工業',
    '2013':'鋼鐵工業','2014':'鋼鐵工業','2015':'鋼鐵工業','2016':'鋼鐵工業',
    '2017':'鋼鐵工業','2018':'鋼鐵工業','2019':'鋼鐵工業','2020':'鋼鐵工業',
    '2022':'鋼鐵工業','2023':'鋼鐵工業','2024':'鋼鐵工業','2025':'鋼鐵工業',
    '2026':'鋼鐵工業','2027':'鋼鐵工業','2028':'鋼鐵工業','2029':'鋼鐵工業',
    '2030':'鋼鐵工業','2031':'鋼鐵工業','2032':'鋼鐵工業','2033':'鋼鐵工業',
    '2034':'鋼鐵工業','2035':'鋼鐵工業','2036':'鋼鐵工業','2038':'鋼鐵工業',
    '2039':'鋼鐵工業','2040':'鋼鐵工業','2041':'鋼鐵工業','2042':'鋼鐵工業',
    '2043':'鋼鐵工業','2044':'鋼鐵工業','2045':'鋼鐵工業','2046':'鋼鐵工業',
    '2047':'鋼鐵工業','2048':'鋼鐵工業','2049':'鋼鐵工業',
    '2062':'鋼鐵工業','2064':'鋼鐵工業','2069':'鋼鐵工業',
    '2072':'鋼鐵工業','2073':'鋼鐵工業','2076':'鋼鐵工業',
    '2083':'鋼鐵工業','2084':'鋼鐵工業','2092':'鋼鐵工業',
    '2093':'鋼鐵工業','2095':'鋼鐵工業','2096':'鋼鐵工業',
    '2304':'光電業','2305':'電腦及週邊設備業','2306':'電腦及週邊設備業',
    '2307':'電腦及週邊設備業','2309':'鋼鐵工業',
    '2059':'其他電子業',
    '2101':'橡膠工業','2102':'橡膠工業','2103':'橡膠工業','2104':'橡膠工業',
    '2105':'橡膠工業','2106':'橡膠工業','2107':'橡膠工業','2108':'橡膠工業',
    '2109':'橡膠工業','2110':'橡膠工業',
    '2201':'汽車工業','2202':'汽車工業','2203':'汽車工業','2204':'汽車工業',
    '2205':'汽車工業','2206':'汽車工業','2207':'汽車工業','2208':'汽車工業',
    '2209':'汽車工業','2210':'汽車工業','2211':'汽車工業','2212':'汽車工業',
    '2213':'汽車工業','2214':'汽車工業','2215':'汽車工業','2216':'汽車工業',
    '2217':'汽車工業','2219':'汽車工業','2220':'汽車工業','2221':'汽車工業',
    '2222':'汽車工業','2223':'汽車工業','2224':'汽車工業','2225':'汽車工業',
    '2226':'汽車工業','2227':'汽車工業','2228':'汽車工業','2229':'汽車工業',
    '2230':'汽車工業','2231':'汽車工業','2232':'汽車工業','2233':'汽車工業',
    '2234':'汽車工業','2235':'汽車工業','2236':'汽車工業','2237':'汽車工業',
    '2239':'汽車工業','2240':'汽車工業','2241':'汽車工業','2243':'汽車工業',
    '2244':'汽車工業','2245':'汽車工業',
    '2301':'電腦及週邊設備業','2303':'半導體業',
    '2308':'光電業',
    '2317':'電腦及週邊設備業',
    '2327':'電子零組件業','2330':'半導體業','2337':'半導體業','2344':'半導體業',
    '2345':'通訊網路業','2347':'半導體業','2348':'電子通路業','2351':'半導體業',
    '2353':'電腦及週邊設備業','2356':'電子通路業','2357':'電腦及週邊設備業',
    '2360':'電子零組件業','2362':'資訊服務業','2365':'電子通路業',
    '2376':'電腦及週邊設備業','2377':'電腦及週邊設備業',
    '2379':'通訊網路業','2382':'電腦及週邊設備業','2383':'電子零組件業',
    '2385':'電腦及週邊設備業','2388':'半導體業','2392':'電腦及週邊設備業',
    '2395':'電腦及週邊設備業',
    '2408':'半導體業','2409':'光電業','2412':'電信服務業','2439':'電子零組件業',
    '2449':'半導體業','2454':'半導體業','2474':'光電業',
    '2501':'建材營造業','2502':'建材營造業','2503':'建材營造業','2504':'建材營造業',
    '2505':'建材營造業','2506':'建材營造業','2507':'建材營造業','2508':'建材營造業',
    '2509':'建材營造業','2511':'建材營造業','2514':'建材營造業','2515':'建材營造業',
    '2516':'建材營造業','2517':'建材營造業','2518':'建材營造業','2520':'建材營造業',
    '2521':'建材營造業','2524':'建材營造業','2527':'建材營造業','2528':'建材營造業',
    '2530':'建材營造業','2534':'建材營造業','2535':'建材營造業','2536':'建材營造業',
    '2537':'建材營造業','2538':'建材營造業','2539':'建材營造業','2540':'建材營造業',
    '2542':'建材營造業','2543':'建材營造業','2545':'建材營造業','2546':'建材營造業',
    '2547':'建材營造業','2548':'建材營造業',
    '2601':'航運業','2602':'航運業','2603':'航運業','2605':'航運業',
    '2606':'航運業','2607':'航運業','2608':'航運業','2609':'航運業',
    '2610':'航運業','2611':'航運業','2612':'航運業','2613':'航運業',
    '2614':'航運業','2615':'航運業','2616':'航運業','2617':'航運業',
    '2618':'航運業','2619':'航運業','2620':'航運業',
    '2701':'觀光餐旅業','2702':'觀光餐旅業','2703':'觀光餐旅業','2704':'觀光餐旅業',
    '2705':'觀光餐旅業','2706':'觀光餐旅業','2707':'觀光餐旅業','2708':'觀光餐旅業',
    '2709':'觀光餐旅業','2711':'觀光餐旅業','2712':'觀光餐旅業','2713':'觀光餐旅業',
    '2714':'觀光餐旅業','2715':'觀光餐旅業','2716':'觀光餐旅業','2717':'觀光餐旅業',
    '2718':'觀光餐旅業','2719':'觀光餐旅業','2720':'觀光餐旅業','2721':'觀光餐旅業',
    '2722':'觀光餐旅業','2723':'觀光餐旅業','2724':'觀光餐旅業','2726':'觀光餐旅業',
    '2727':'觀光餐旅業','2728':'觀光餐旅業',
    '2801':'金融保險業','2809':'金融保險業','2812':'金融保險業','2820':'金融保險業',
    '2822':'金融保險業','2823':'金融保險業','2824':'金融保險業','2825':'金融保險業',
    '2826':'金融保險業','2832':'金融保險業','2834':'金融保險業','2836':'金融保險業',
    '2838':'金融保險業','2841':'金融保險業','2845':'金融保險業','2847':'金融保險業',
    '2849':'金融保險業','2850':'金融保險業','2851':'金融保險業','2852':'金融保險業',
    '2855':'金融保險業','2856':'金融保險業','2857':'金融保險業','2860':'金融保險業',
    '2861':'金融保險業','2867':'金融保險業','2868':'金融保險業','2869':'金融保險業',
    '2870':'金融保險業','2871':'金融保險業','2872':'金融保險業','2873':'金融保險業',
    '2874':'金融保險業','2876':'金融保險業','2877':'金融保險業','2878':'金融保險業',
    '2879':'金融保險業','2880':'金融保險業','2881':'金融保險業','2882':'金融保險業',
    '2883':'金融保險業','2884':'金融保險業','2885':'金融保險業','2886':'金融保險業',
    '2887':'金融保險業','2888':'金融保險業','2889':'金融保險業','2890':'金融保險業',
    '2891':'金融保險業','2892':'金融保險業','2893':'金融保險業','2897':'金融保險業',
    '2898':'金融保險業',
    '2901':'貿易百貨業','2903':'貿易百貨業','2904':'貿易百貨業','2906':'貿易百貨業',
    '2908':'貿易百貨業','2910':'貿易百貨業','2911':'貿易百貨業','2912':'貿易百貨業',
    '2913':'貿易百貨業','2915':'貿易百貨業','2916':'貿易百貨業','2917':'貿易百貨業',
    '2918':'貿易百貨業','2919':'貿易百貨業','2920':'貿易百貨業',
    '3006':'半導體業','3034':'半導體業','3037':'半導體業',
    '3045':'電信服務業','3094':'半導體業','3130':'資訊服務業',
    '3231':'電腦及週邊設備業','3260':'電子零組件業','3293':'資訊服務業',
    '3443':'半導體業','3481':'光電業','3529':'半導體業',
    '3552':'電子零組件業','3661':'半導體業','3702':'電子通路業','3711':'電子零組件業',
    '4105':'生技醫療業','4106':'生技醫療業','4107':'生技醫療業','4108':'生技醫療業',
    '4109':'生技醫療業','4110':'生技醫療業','4111':'生技醫療業','4112':'生技醫療業',
    '4113':'生技醫療業','4114':'生技醫療業','4115':'生技醫療業','4116':'生技醫療業',
    '4117':'生技醫療業','4119':'生技醫療業','4121':'生技醫療業','4122':'生技醫療業',
    '4123':'生技醫療業','4124':'生技醫療業','4125':'生技醫療業','4126':'生技醫療業',
    '4127':'生技醫療業','4128':'生技醫療業','4129':'生技醫療業','4130':'生技醫療業',
    '4131':'生技醫療業','4132':'生技醫療業','4133':'生技醫療業','4134':'生技醫療業',
    '4135':'生技醫療業','4136':'生技醫療業','4137':'生技醫療業','4138':'生技醫療業',
    '4139':'生技醫療業','4140':'生技醫療業','4141':'生技醫療業','4142':'生技醫療業',
    '4144':'生技醫療業','4148':'生技醫療業','4152':'生技醫療業','4154':'生技醫療業',
    '4155':'生技醫療業','4157':'生技醫療業','4158':'生技醫療業','4159':'生技醫療業',
    '4160':'生技醫療業','4161':'生技醫療業','4162':'生技醫療業','4163':'生技醫療業',
    '4164':'生技醫療業','4165':'生技醫療業','4166':'生技醫療業','4167':'生技醫療業',
    '4168':'生技醫療業','4169':'生技醫療業','4170':'生技醫療業','4171':'生技醫療業',
    '4172':'生技醫療業','4173':'生技醫療業','4174':'生技醫療業','4175':'生技醫療業',
    '4176':'生技醫療業','4177':'生技醫療業','4179':'生技醫療業','4180':'生技醫療業',
    '4183':'生技醫療業','4188':'生技醫療業','4192':'生技醫療業','4194':'生技醫療業',
    '4198':'生技醫療業','4200':'生技醫療業','4205':'生技醫療業','4206':'生技醫療業',
    '4207':'生技醫療業','4209':'生技醫療業','4210':'生技醫療業','4211':'生技醫療業',
    '4215':'生技醫療業','4217':'生技醫療業','4218':'生技醫療業','4219':'生技醫療業',
    '4220':'生技醫療業','4221':'生技醫療業','4222':'生技醫療業','4223':'生技醫療業',
    '4224':'生技醫療業','4225':'生技醫療業','4226':'生技醫療業','4227':'生技醫療業',
    '4228':'生技醫療業','4229':'生技醫療業','4230':'生技醫療業','4232':'生技醫療業',
    '4904':'電信服務業','4919':'半導體業','4938':'電腦及週邊設備業',
    '4958':'通訊網路業','4966':'半導體業',
    '5274':'半導體業','5425':'半導體業','5483':'半導體業',
    '5522':'建材營造業','5534':'建材營造業',
    '5820':'金融保險業','5821':'金融保險業','5823':'金融保險業','5824':'金融保險業',
    '5825':'金融保險業','5826':'金融保險業','5827':'金融保險業','5830':'金融保險業',
    '5832':'金融保險業','5834':'金融保險業','5835':'金融保險業','5836':'金融保險業',
    '5838':'金融保險業','5843':'金融保險業','5844':'金融保險業','5845':'金融保險業',
    '5846':'金融保險業','5847':'金融保險業','5850':'金融保險業','5853':'金融保險業',
    '5854':'金融保險業','5855':'金融保險業','5856':'金融保險業','5857':'金融保險業',
    '5859':'金融保險業','5860':'金融保險業','5861':'金融保險業','5863':'金融保險業',
    '5864':'金融保險業','5865':'金融保險業','5866':'金融保險業','5867':'金融保險業',
    '5868':'金融保險業','5869':'金融保險業','5871':'金融保險業',
    '6176':'光電業','6197':'電子零組件業','6274':'電子零組件業',
    '6278':'電子零組件業','6414':'通訊網路業','6415':'半導體業',
    '6488':'半導體業','6505':'油電燃氣業','6669':'電腦及週邊設備業',
    '6770':'半導體業',
    '8046':'半導體業','8103':'資訊服務業',
    '9910':'橡膠工業','9945':'建材營造業','9951':'油電燃氣業',
    '9955':'油電燃氣業','9957':'油電燃氣業','9958':'油電燃氣業',
}

def get_industry_from_api():
    """
    從 TWSE OpenAPI 動態抓取產業別，僅補充靜態表未涵蓋的股票。
    靜態 INDUSTRY_MAP 為主要來源，此函式只負責補漏。
    """
    # 合法的中文產業名稱白名單（防止 API 亂碼或錯誤欄位混入）
    VALID_NAMES = set(INDUSTRY_CODE_TO_NAME.values())

    industry = {}
    hdrs = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'application/json',
        'Referer': 'https://www.twse.com.tw/',
    }

    for label, url in [
        ('上市', 'https://openapi.twse.com.tw/v1/opendata/t187ap03_L'),
        ('上櫃', 'https://openapi.twse.com.tw/v1/opendata/t187ap03_O'),
    ]:
        try:
            resp = requests.get(url, timeout=15, headers=hdrs)
            if resp.status_code != 200:
                print(f'  [{label}] API HTTP {resp.status_code}，跳過')
                continue

            data = resp.json()
            if not data:
                continue

            # ── 診斷：印出第一筆的所有欄位，方便確認 ──────────────
            sample = data[0]
            print(f'  [{label}] API 欄位: {list(sample.keys())}')
            print(f'  [{label}] 第一筆範例: {sample}')

            # ── 偵測「股票代號」欄 ─────────────────────────────────
            code_col = next(
                (f for f in ['公司代號', '股票代號', 'Code', 'StockCode']
                 if f in sample), None
            )
            # ── 偵測「產業別」欄 ───────────────────────────────────
            ind_col = next(
                (f for f in ['產業別', '產業類別', '產業別名稱',
                             '產業類別代號', 'IndustryName', 'industry']
                 if f in sample), None
            )

            if not code_col or not ind_col:
                print(f'  [{label}] 找不到必要欄位 (code={code_col}, ind={ind_col})，跳過')
                continue

            ok = 0
            for item in data:
                code    = str(item.get(code_col, '')).strip()
                raw_ind = str(item.get(ind_col, '')).strip()

                # 只處理 4 位數字股票代碼
                if not (code.isdigit() and len(code) == 4):
                    continue
                if not raw_ind or raw_ind in ('nan', 'None', '', '-'):
                    continue

                # 若 API 回傳數字代號 → 用對照表轉中文
                if raw_ind.isdigit():
                    mapped = INDUSTRY_CODE_TO_NAME.get(raw_ind.zfill(2), '')
                    if not mapped:
                        continue          # 未知代號，略過（不寫入）
                    raw_ind = mapped

                # 白名單驗證：只接受合法的產業中文名稱
                if raw_ind not in VALID_NAMES:
                    continue              # 非預期值（亂碼/英文/異常），略過

                # 靜態表已有的股票不覆蓋（靜態表是人工驗證過的）
                if code not in INDUSTRY_MAP:
                    industry[code] = raw_ind
                    ok += 1

            print(f'  [{label}] 補充了 {ok} 筆靜態表未涵蓋的股票產業別')

        except Exception as e:
            print(f'  [{label}] API 例外：{e}')

    total = len(industry)
    if total:
        print(f'✅ API 額外補充：{total} 筆（靜態表以外的股票）')
    else:
        print('ℹ️  API 未補充新筆數（靜態表已全涵蓋，或 API 不可用）')
    return industry


def lookup_industry(code, isin_industry, static_map=None):
    """
    優先順序：① ISIN HTML（TWSE 官方，每次執行即時抓取）
              ② 靜態 INDUSTRY_MAP（人工驗證備援）
              ③ '其他'
    """
    # ① ISIN 即時資料（最可靠）
    if code in isin_industry:
        val = isin_industry[code]
        if val and val not in ('nan', 'NaN', 'None', '', '-'):
            return val
    # ② 靜態備援
    sm = static_map if static_map is not None else INDUSTRY_MAP
    if code in sm:
        return sm[code]
    return '其他'


# ===============================================================
# 3. 技術指標
# ===============================================================
# 這兩支本來是 main() 裡的內部函式。搬到模組層唯一的理由是：能被測到。
# 篩選的四道關卡全都建立在它們算出來的數字上，而一個算錯的布林帶不會讓程式
# 當掉，只會讓每天的名單默默地變成另一份名單。
#
# 公式一個字都沒有動。

def compute_atr(df, period=14):
    """Wilder 的真實波幅，用簡單移動平均而不是指數平滑——和本機版一致。"""
    h, l, c = df['High'], df['Low'], df['Close']
    tr = pd.concat([(h-l),(h-c.shift()).abs(),(l-c.shift()).abs()], axis=1).max(axis=1)
    return tr.rolling(period).mean()


def compute_bollinger(close, period=20, k=2):
    """回傳 ``(中軌, 上軌, 下軌, 頻寬)``。頻寬 = (上－下) / 中，用來判斷壓縮。"""
    ma  = close.rolling(period).mean()
    std = close.rolling(period).std()
    up  = ma + k*std
    dn  = ma - k*std
    bw  = (up - dn) / ma
    return ma, up, dn, bw


def run(
    output_dir: str,
    *,
    limit: int = 0,
    workers: int = 8,
    period: str = '2y',
    make_excel: bool = True,
    excel_charts: bool = True,
    link_base: str = '',
    plotly_cdn: bool = True,
    chart_years: float = 2.0,
    excel_url: str = '',
    index_copy: str = '',
    open_when_done: bool = False,
) -> dict:
    """跑完一次全市場篩選，回傳 ``{'results', 'xlsx', 'html', 'index'}``。

    參數只有三種：**跑多少**（``limit`` / ``workers`` / ``period``）、**產出什麼**
    （``make_excel`` / ``excel_charts`` / ``index_copy``）、**報告長什麼樣**
    （``link_base`` / ``plotly_cdn`` / ``chart_years`` / ``excel_url``）。篩選規則
    本身沒有參數——那是這支程式的定義，不是設定。

    ``period='2y'`` 而不是本機版的 ``'5y'``：所有指標裡窗口最長的是 60MA 加上
    Donchian 的 20 日位移，八十幾個交易日。兩年和五年算出來的最後一列一模一樣，
    但要下載的資料少六成——乘上一千九百檔，那是排程跑不跑得完的差別。想完整
    重現本機版的下載行為就傳 ``period='5y'``。
    """
    print('=' * 60)
    print(f'  台股順勢交易篩選系統 {VERSION}')
    print('=' * 60)
    print('[1/3] 載入上市＋上櫃股票清單（含產業別）...')
    TICKERS, NAME_MAP, ISIN_INDUSTRY = load_tw_stock_universe()
    if limit:
        # 只留前 N 檔。給 CI 的煙霧測試用：它要證明的是「整條路走得通」，不是
        # 「今天有哪幾檔突破」，而後者要跑二十分鐘。
        TICKERS = TICKERS[:limit]
        print(f'   （--limit {limit}：只掃前 {len(TICKERS)} 檔）')
    print(f'✅ 股票池：{len(TICKERS)} 檔  |  ISIN 產業別：{len(ISIN_INDUSTRY)} 筆')
    print()

    # ===============================================================
    # 3. 技術指標計算 → 移到模組層（見 compute_atr / compute_bollinger）
    # ===============================================================
    def screen_stock(ticker):
        try:
            try:
                df = yf.download(ticker, period=period, interval='1d',
                                 auto_adjust=True, progress=False, threads=False,
                                 session=_YF_SESSION)
            except Exception:
                df = None
            if df is None or len(df) < 65:
                # 備援：某些環境下自訂 session 反而會失敗，改用預設連線再試一次
                try:
                    df = yf.download(ticker, period=period, interval='1d',
                                     auto_adjust=True, progress=False, threads=False)
                except Exception:
                    df = None
            if df is None or len(df) < 65: return None
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.droplevel(1)
            df = df.dropna(subset=['Close','Volume'])
            df.index = pd.to_datetime(df.index).tz_localize(None)
            close, volume = df['Close'], df['Volume']

            price = float(close.iloc[-1])
            vol20 = float(volume.rolling(20).mean().iloc[-1])
            amt20 = float((close*volume).rolling(20).mean().iloc[-1])
            if price <= 10 or vol20 <= 1000 or amt20 <= 50_000_000: return None

            ma20 = close.rolling(20).mean()
            ma60 = close.rolling(60).mean()
            if price <= float(ma60.iloc[-1]) or float(ma20.iloc[-1]) <= float(ma60.iloc[-1]):
                return None

            boll_ma, boll_up, boll_dn, bw = compute_bollinger(close)
            lookback = min(10, len(ma20)-1)
            golden_cross = any(
                float(ma20.iloc[-i]) > float(ma60.iloc[-i]) and
                float(ma20.iloc[-i-1]) <= float(ma60.iloc[-i-1])
                for i in range(1, lookback+1)
            )
            squeeze = bool((bw.iloc[-lookback:] <= 0.12).any())
            if not (golden_cross or squeeze): return None

            donchian     = close.rolling(20).max().shift(1)
            brk_boll     = price > float(boll_up.iloc[-1])
            brk_donchian = (not pd.isna(donchian.iloc[-1])) and (price > float(donchian.iloc[-1]))
            if not (brk_boll or brk_donchian): return None

            vol_ratio = float(volume.iloc[-1]) / vol20
            if vol_ratio < 1.2: return None

            atr_val   = float(compute_atr(df).iloc[-1])
            stop_loss = round(price - 3*atr_val, 2)

            trigger_parts = []
            if golden_cross:   trigger_parts.append('黃金交叉')
            if squeeze:        trigger_parts.append('布林壓縮')
            if brk_boll:       trigger_parts.append('突破布林上軌')
            if brk_donchian:   trigger_parts.append('突破20日高點')

            code = ticker.split('.')[0]
            return {
                'ticker': ticker, 'code': code, 'name': NAME_MAP.get(code, code),
                'industry': lookup_industry(code, ISIN_INDUSTRY),
                'close':         round(price, 2),
                'ma20_last':     round(float(ma20.iloc[-1]), 2),
                'ma60_last':     round(float(ma60.iloc[-1]), 2),
                'boll_up_last':  round(float(boll_up.iloc[-1]), 2),
                'boll_mid_last': round(float(boll_ma.iloc[-1]), 2),
                'boll_dn_last':  round(float(boll_dn.iloc[-1]), 2),
                'boll_bw_pct':   round(float(bw.iloc[-1])*100, 2),
                'vol_today':     int(volume.iloc[-1]),
                'vol20_avg':     round(vol20, 0),
                'vol_ratio':     round(vol_ratio, 2),
                'amt_M':         round(amt20/1e6, 1),
                'atr14':         round(atr_val, 2),
                'stop_loss':     stop_loss,
                'golden_cross':  golden_cross,
                'squeeze':       squeeze,
                'trigger':       ' ｜ '.join(trigger_parts),
                '_df':           df.copy(),
                '_ma20':         ma20.copy(),
                '_ma60':         ma60.copy(),
                '_boll_up':      boll_up.copy(),
                '_boll_mid':     boll_ma.copy(),
                '_boll_dn':      boll_dn.copy(),
            }
        except Exception:
            return None

    # ===============================================================
    # 4. 高質感 K 線 + 量能圖
    # ===============================================================
    def draw_stock_chart(res, img_path):
        df = res['_df'].copy()
        n  = min(60, len(df))
        df = df.iloc[-n:]

        ma20   = res['_ma20'].iloc[-n:]
        ma60   = res['_ma60'].iloc[-n:]
        bup    = res['_boll_up'].iloc[-n:]
        bmid   = res['_boll_mid'].iloc[-n:]
        bdn    = res['_boll_dn'].iloc[-n:]
        volume = df['Volume']
        dates  = np.arange(len(df))

        C = COLORS
        plt.rcParams['path.simplify']           = True
        plt.rcParams['path.simplify_threshold'] = 0.5

        fig = plt.figure(figsize=(16, 9), facecolor=C['bg'])
        gs  = gridspec.GridSpec(
            2, 1,
            height_ratios=[3, 1],
            hspace=0.10,
            top=0.88,
            bottom=0.11,
            left=0.07,
            right=0.96
        )
        ax1 = fig.add_subplot(gs[0]); ax1.set_facecolor(C['axes_bg'])
        ax2 = fig.add_subplot(gs[1], sharex=ax1); ax2.set_facecolor(C['axes_bg'])

        # 布林通道填色
        ax1.fill_between(dates, bup.values, bdn.values,
                         alpha=0.10, color='#7ee787', zorder=1)

        # K 棒
        for i, (_, row) in enumerate(df.iterrows()):
            o, h, l, cp = row['Open'], row['High'], row['Low'], row['Close']
            col = C['vol_up'] if cp >= o else C['vol_dn']
            ax1.plot([i, i], [l, h], color=col, linewidth=1.0, zorder=3)
            ax1.add_patch(Rectangle((i-0.38, min(o, cp)), 0.76, abs(cp-o),
                                      color=col, zorder=4))

        # 均線＋布林線
        ax1.plot(dates, ma20.values, color=C['ma20'], lw=2.2,
                 label='20MA 月線', zorder=5)
        ax1.plot(dates, ma60.values, color=C['ma60'], lw=2.4, ls='--',
                 label='60MA 季線', zorder=5)
        ax1.plot(dates, bup.values,  color='#7ee787', lw=1.2, ls='--',
                 label='布林上軌', zorder=5)
        ax1.plot(dates, bmid.values, color='#8b949e', lw=1.0, ls=':',
                 label='布林中軌', zorder=5)
        ax1.plot(dates, bdn.values,  color='#7ee787', lw=1.2, ls='--',
                 label='布林下軌', zorder=5)

        # 停損線
        ax1.axhline(y=res['stop_loss'], color=C['stop'], lw=1.5, ls='-.',
                    label=f"建議停損 {res['stop_loss']:.2f}", zorder=6)

        # 最新收盤標記
        ax1.annotate(f"  ▶ {res['close']:.2f}",
                     xy=(len(dates)-1, res['close']),
                     color='#58a6ff', fontsize=11, va='center',
                     fontproperties=FP, zorder=7)

        ax1.set_xlim(-1, len(dates))
        ax1.grid(color=C['grid'], lw=0.5, alpha=0.6)
        ax1.tick_params(colors=C['text'], labelsize=10)
        for sp in ax1.spines.values(): sp.set_color(C['grid'])
        plt.setp(ax1.get_xticklabels(), visible=False)

        ax1.legend(
            loc='upper left', fontsize=9.5, framealpha=0.35,
            facecolor=C['axes_bg'], edgecolor=C['grid'],
            labelcolor=C['text'], prop=FP, ncol=3
        )

        # 成交量
        vol20_ma = volume.rolling(20).mean()
        vcols = [C['vol_up'] if df['Close'].iloc[i] >= df['Open'].iloc[i]
                 else C['vol_dn'] for i in range(len(df))]
        ax2.bar(dates, volume.values, color=vcols, width=0.65, alpha=0.85, zorder=3)
        ax2.plot(dates, vol20_ma.values, color='#ffa500', lw=1.5, zorder=4, label='20日均量')
        ax2.set_xlim(-1, len(dates))
        ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{x/1e6:.1f}M'))
        ax2.grid(color=C['grid'], lw=0.4, alpha=0.5)
        ax2.tick_params(colors=C['text'], labelsize=9)
        for sp in ax2.spines.values(): sp.set_color(C['grid'])

        # 日期標籤（量圖底部，不被遮蓋）
        step = max(1, len(dates) // 10)
        tick_pos    = list(range(0, len(dates), step))
        tick_labels = [df.index[i].strftime('%m/%d') for i in tick_pos]
        ax2.set_xticks(tick_pos)
        ax2.set_xticklabels(tick_labels, color=C['text'], fontsize=9.5,
                            fontproperties=FP, rotation=0)
        ax2.tick_params(axis='x', pad=6)

        ax2.legend(
            loc='upper right', fontsize=9, framealpha=0.35,
            facecolor=C['axes_bg'], edgecolor=C['grid'],
            labelcolor=C['text'], prop=FP
        )

        # 圖表標題文字
        t1 = (f"{res['code']}  {res['name']}    "
              f"收盤：{res['close']:.2f}    "
              f"20MA：{res['ma20_last']:.2f}    "
              f"60MA：{res['ma60_last']:.2f}    "
              f"ATR(14)：{res['atr14']:.2f}    "
              f"建議停損：{res['stop_loss']:.2f}")
        t2 = (f"觸發訊號：{res['trigger']}    ｜    "
              f"量比：{res['vol_ratio']:.2f}x    ｜    "
              f"布林頻寬：{res['boll_bw_pct']:.1f}%    ｜    "
              f"成交金額：{res['amt_M']:.1f} 百萬")

        fig.text(0.50, 0.958, t1, ha='center', va='top',
                 fontsize=12.5, color='#e6edf3', fontproperties=FP_BOLD)
        fig.text(0.50, 0.922, t2, ha='center', va='top',
                 fontsize=10.5, color='#ffa500', fontproperties=FP)

        plt.savefig(img_path, dpi=150, bbox_inches='tight', facecolor=C['bg'])
        plt.close()

    # ===============================================================
    # 5. 執行全市場篩選
    # ===============================================================
    TW_TZ     = datetime.timezone(datetime.timedelta(hours=8))
    now       = datetime.datetime.now(TW_TZ)
    today_str = now.strftime('%Y-%m-%d')

    print('='*60)
    print(f'📊 [2/3] 台股順勢交易系統 {VERSION} ── 嚴格四部曲篩選')
    print('='*60)
    print(f'掃描股票池：{len(TICKERS)} 檔  ｜  日期：{today_str}  ｜  歷史長度：{period}')
    print('-'*60)

    from concurrent.futures import ThreadPoolExecutor
    import threading

    _lock    = threading.Lock()
    RESULTS  = []
    _scanned = [0]

    def _screen_and_collect(args):
        i, ticker = args
        r = screen_stock(ticker)
        with _lock:
            _scanned[0] += 1
            if r:
                RESULTS.append(r)
                print(f"✅ {r['code']:>4s} {r['name']:<8s} "
                      f"收:{r['close']:>7.2f} 量比:{r['vol_ratio']:.2f}x │ {r['trigger']}")
            if _scanned[0] % 100 == 0:
                print(f'── 已掃描 {_scanned[0]}/{len(TICKERS)} 檔 ──')

    with ThreadPoolExecutor(max_workers=workers) as executor:
        list(executor.map(_screen_and_collect, enumerate(TICKERS)))

    RESULTS.sort(key=lambda x: x['vol_ratio'], reverse=True)
    print('='*60)
    print(f'🎯 篩選完成！今日共 {len(RESULTS)} 檔標的通過')
    print('='*60)

    # ===============================================================
    # 6. 建立 Excel（3 分頁）
    # ===============================================================
    os.makedirs(output_dir, exist_ok=True)
    OUTPUT_FILE = ''
    if make_excel:
        IMG_DIR     = tempfile.mkdtemp()
        OUTPUT_FILE = os.path.join(output_dir, now.strftime(
            f'TW_Stock_Trend_Following_Trading_Filter_Result_{VERSION}_%Y_%m_%d_%H%M.xlsx'))
        wb          = Workbook()

        # ──────────────────────────────────────────────
        # Sheet 1：篩選總覽
        # ──────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = '篩選總覽'
        ws1.column_dimensions['A'].width = 2
        # 新增 K 欄：原 B~J 不變，新增 K 欄（產業別）
        for col, w in zip('BCDEFGHIJK', [18.5, 18, 12, 10, 48.5, 10, 12, 30, 9, 11]):
            ws1.column_dimensions[col].width = w

        ws1.merge_cells('B2:K2')
        c = ws1['B2']
        c.value     = '📊 台股順勢交易系統 V3.1 ── 嚴格四部曲篩選報告'
        c.font      = mkfont('Calibri', 18, True, 'FFFFFF')
        c.fill      = mkfill(C_NAVY)
        c.alignment = ALN_CC
        ws1.row_dimensions[2].height = 38

        ws1.merge_cells('B3:K3')
        c = ws1['B3']
        c.value     = (f'篩選日期：{today_str}    掃描股票池：{len(TICKERS)} 檔    '
                       f'今日通過篩選：{len(RESULTS)} 檔標的')
        c.font      = mkfont('Calibri', 11, False, 'FFFFFF')
        c.fill      = mkfill(C_TEAL)
        c.alignment = ALN_CC
        ws1.row_dimensions[3].height = 22
        ws1.row_dimensions[4].height = 6

        ws1.merge_cells('B5:K5')
        c = ws1['B5']
        c.value     = '【四部曲篩選機制說明】'
        c.font      = mkfont('Calibri', 12, True, 'FFFFFF')
        c.fill      = mkfill(C_HEADER)
        c.alignment = ALN_LC
        ws1.row_dimensions[5].height = 22

        strat = [
            ('① 基礎流動性防禦', '股價 > 10 元 ｜ 20日均量 > 1,000 張 ｜ 日均成交金額 > 5,000 萬元'),
            ('② 趨勢多頭確認',   '收盤站穩季線(60MA)之上，且月線(20MA) > 季線(60MA)'),
            ('③ 關鍵發動時機',   '過去 10 日內：月季線黃金交叉 或 布林頻寬壓縮 ≤ 12%'),
            ('④ 強勢突破＋爆量', '收盤突破布林上軌 或 創 20 日新高，且當日量 ≥ 20 日均量 × 1.2'),
        ]
        for i, (lbl, desc) in enumerate(strat, start=6):
            ws1.row_dimensions[i].height = 22
            c_l = ws1.cell(i, 2, lbl)
            c_l.font      = mkfont('Calibri', 10, True, '1A3A5C')
            c_l.fill      = mkfill('EBF5FB')
            c_l.alignment = ALN_LCW
            c_l.border    = border_all()
            ws1.merge_cells(f'C{i}:K{i}')
            c_d = ws1.cell(i, 3, desc)
            c_d.font      = mkfont('Calibri', 10, False, '2C3E50')
            c_d.fill      = mkfill('FDFEFE')
            c_d.alignment = ALN_LCW
            c_d.border    = border_all()

        ws1.row_dimensions[10].height = 6

        ws1.merge_cells('B11:K11')
        c = ws1['B11']
        c.value     = '【進出場策略】'
        c.font      = mkfont('Calibri', 12, True, 'FFFFFF')
        c.fill      = mkfill(C_HEADER)
        c.alignment = ALN_LC
        ws1.row_dimensions[11].height = 22

        rules = [
            ('進場時機', '訊號觸發後次一交易日，開盤直接以市價單敲進（切勿掛低價等待）。'),
            ('初始停損', '停損 = 最新收盤價 − 3 × ATR(14)，進場後立即設定並固定不放寬。'),
            ('動態移停', '每日收盤後，可將停損單往上調整至當日最新 20MA 附近。'),
            ('終極出場', '當收盤價正式跌破當日 20MA 時，考慮全數出場鎖定波段獲利。'),
        ]
        for i, (lbl, desc) in enumerate(rules, start=12):
            ws1.row_dimensions[i].height = 22
            c_l = ws1.cell(i, 2, lbl)
            c_l.font      = mkfont('Calibri', 10, True, '6E2F00')
            c_l.fill      = mkfill('FEF9E7')
            c_l.alignment = ALN_LCW
            c_l.border    = border_all()
            ws1.merge_cells(f'C{i}:K{i}')
            c_d = ws1.cell(i, 3, desc)
            c_d.font      = mkfont('Calibri', 10, False, '2C3E50')
            c_d.fill      = mkfill('FFFFF0')
            c_d.alignment = ALN_LCW
            c_d.border    = border_all()

        ws1.row_dimensions[16].height = 6

        ws1.merge_cells('B17:K17')
        c = ws1['B17']
        c.value     = f'【篩選結果摘要】今日共篩選出 {len(RESULTS)} 檔標的通過'
        c.font      = mkfont('Calibri', 12, True, 'FFFFFF')
        c.fill      = mkfill('1A5276')
        c.alignment = ALN_LC
        ws1.row_dimensions[17].height = 22

        # 欄順序：代號(B) 名稱(C) 產業別(D) 收盤價(E) 觸發訊號(F) 20MA(G) 布林上軌(H) 量比(I) 建議停損(J) 60MA(K)
        SHDR = ['代號','名稱','產業別','收盤價','觸發訊號','20MA','布林上軌','量比(倍)','建議停損','60MA']
        ws1.row_dimensions[18].height = 22
        for ci, h in enumerate(SHDR, start=2):
            c = ws1.cell(18, ci, h)
            c.font      = mkfont('Calibri', 10, True, 'FFFFFF')
            c.fill      = mkfill(C_HEADER)
            c.alignment = ALN_CCW
            c.border    = border_all(False)

        for ri, res in enumerate(RESULTS, start=19):
            ws1.row_dimensions[ri].height = 20
            fill = mkfill(C_ROW1) if ri % 2 == 1 else mkfill(C_ROW2)
            row_vals = [
                res['code'],        # B  代號
                res['name'],        # C  名稱
                res['industry'],    # D  產業別  ← 新增
                res['close'],       # E  收盤價
                res['trigger'],     # F  觸發訊號
                res['ma20_last'],   # G  20MA
                res['boll_up_last'],# H  布林上軌
                res['vol_ratio'],   # I  量比
                res['stop_loss'],   # J  建議停損
                res['ma60_last'],   # K  60MA
            ]
            for ci, val in enumerate(row_vals, start=2):
                c = ws1.cell(ri, ci, val)
                c.fill   = fill
                c.border = border_all()
                if ci == 4:   # 產業別 D
                    c.font      = mkfont('Calibri', 10, True, '154360')
                    c.fill      = mkfill('D6EAF8')
                    c.alignment = ALN_CC
                elif ci == 5:   # 觸發訊號 F（原 ci==4）
                    c.font      = mkfont('Calibri', 10, True, '6E2F00')
                    c.alignment = ALN_LCW
                elif ci in (6, 7, 11):  # 20MA / 布林上軌 / 60MA
                    c.font          = mkfont('Calibri', 10, False, '2C3E50')
                    c.alignment     = ALN_RC
                    c.number_format = '#,##0.00'
                elif ci == 9:   # 量比 I
                    ratio = float(val)
                    fg = 'C0392B' if ratio >= 2.0 else ('196F3D' if ratio >= 1.5 else '2C3E50')
                    c.font          = mkfont('Calibri', 10, True, fg)
                    c.alignment     = ALN_CC
                    c.number_format = '0.00"x"'
                elif ci == 10:  # 建議停損 J
                    c.font          = mkfont('Calibri', 10, True, C_ACCENT)
                    c.fill          = mkfill(C_STOPBG)
                    c.alignment     = ALN_CC
                    c.number_format = '#,##0.00'
                else:
                    c.font      = mkfont('Calibri', 10, bold=(ci == 2), color='2C3E50')
                    c.alignment = ALN_CC

        if RESULTS:
            last_s = 18 + len(RESULTS)
            tbl = Table(displayName='SummaryTable', ref=f'B18:K{last_s}')
            tbl.tableStyleInfo = TableStyleInfo(name='TableStyleMedium9', showRowStripes=True)
            ws1.add_table(tbl)

        # 產業別欄(D) 依實際內容自動調整最佳欄寬
        max_ind_len_ws1 = max(
            (len(str(res['industry'])) for res in RESULTS),
            default=4
        )
        # 中文字約佔 2 個英文字元寬度，加 4 為邊距
        ws1.column_dimensions['D'].width = max(10, min(max_ind_len_ws1 * 2 + 4, 30))

        # ──────────────────────────────────────────────
        # Sheet 2：詳細數據
        # ──────────────────────────────────────────────
        ws2 = wb.create_sheet('詳細數據')
        ws2.column_dimensions['A'].width = 2
        # 在「名稱(C)」後插入「產業別(D)」，原 D 以後順移一欄至 E~T
        COL_DEFS = [
            ('B',   8,    '代號'),
            ('C',  14,    '名稱'),
            ('D',  12,    '產業別'),   # ← 新增
            ('E',  10,    '收盤價'),
            ('F',  10,    '20MA'),
            ('G',  10,    '60MA'),
            ('H',  10,    '布林上軌'),
            ('I',  10,    '布林中軌'),
            ('J',  10,    '布林下軌'),
            ('K',  10,    '布林頻寬%'),
            ('L',  14,    '今日成交量'),
            ('M',  12,    '20日均量'),
            ('N',   9,    '量比(倍)'),
            ('O',  14,    '均成交額(百萬)'),
            ('P',   9,    'ATR(14)'),
            ('Q',  12,    '建議停損'),
            ('R',   9,    '黃金交叉'),
            ('S',   9,    '布林壓縮'),
            ('T',  47.5,  '觸發訊號'),
        ]
        for col, w, _ in COL_DEFS:
            ws2.column_dimensions[col].width = w

        ws2.merge_cells('B2:T2')
        c = ws2['B2']
        c.value     = f'台股順勢交易詳細數據 V3.1 ｜ 篩選日期：{today_str} ｜ 共 {len(RESULTS)} 檔標的'
        c.font      = mkfont('Calibri', 13, True, 'FFFFFF')
        c.fill      = mkfill(C_NAVY)
        c.alignment = ALN_CC
        ws2.row_dimensions[2].height = 30

        ws2.row_dimensions[3].height = 22
        for ci, (_, _, hdr) in enumerate(COL_DEFS, start=2):
            c = ws2.cell(3, ci, hdr)
            c.font      = mkfont('Calibri', 10, True, 'FFFFFF')
            c.fill      = mkfill(C_HEADER)
            c.alignment = ALN_CCW
            c.border    = border_all(False)

        for ri, res in enumerate(RESULTS, start=4):
            ws2.row_dimensions[ri].height = 20
            fill = mkfill(C_ROW1) if ri % 2 == 0 else mkfill(C_ROW2)
            # 欄位順序對應 COL_DEFS: B=代號 C=名稱 D=產業別 E=收盤 F=20MA G=60MA
            # H=布林上 I=布林中 J=布林下 K=布林頻寬 L=今日量 M=均量 N=量比
            # O=均成交額 P=ATR Q=停損 R=黃金交叉 S=布林壓縮 T=觸發訊號
            row_vals = [
                res['code'],                              # B  ci=2
                res['name'],                              # C  ci=3
                res['industry'],                          # D  ci=4  ← 新增
                res['close'],                             # E  ci=5
                res['ma20_last'],                         # F  ci=6
                res['ma60_last'],                         # G  ci=7
                res['boll_up_last'],                      # H  ci=8
                res['boll_mid_last'],                     # I  ci=9
                res['boll_dn_last'],                      # J  ci=10
                res['boll_bw_pct'],                       # K  ci=11
                res['vol_today'],                         # L  ci=12
                res['vol20_avg'],                         # M  ci=13
                res['vol_ratio'],                         # N  ci=14
                res['amt_M'],                             # O  ci=15
                res['atr14'],                             # P  ci=16
                res['stop_loss'],                         # Q  ci=17
                '✅' if res['golden_cross'] else '—',    # R  ci=18
                '✅' if res['squeeze']      else '—',    # S  ci=19
                res['trigger'],                           # T  ci=20
            ]
            for ci, val in enumerate(row_vals, start=2):
                c = ws2.cell(ri, ci, val)
                c.fill      = fill
                c.border    = border_all()
                c.alignment = ALN_CC
                if ci == 4:   # 產業別 D
                    c.font      = mkfont('Calibri', 10, True, '154360')
                    c.fill      = mkfill('D6EAF8')
                    c.alignment = ALN_CC
                elif ci in (5, 6, 7, 8, 9, 10, 16, 17):  # 數值欄（收盤/均線/布林/ATR/停損）
                    c.number_format = '#,##0.00'
                    c.font = mkfont('Calibri', 10, False, '2C3E50')
                elif ci in (12, 13):  # 今日量 / 均量
                    c.number_format = '#,##0'
                    c.font = mkfont('Calibri', 10, False, '2C3E50')
                elif ci == 14:  # 量比 N
                    ratio = float(val)
                    fg = 'C0392B' if ratio >= 2 else ('196F3D' if ratio >= 1.5 else '2C3E50')
                    c.font          = mkfont('Calibri', 10, True, fg)
                    c.number_format = '0.00"x"'
                elif ci == 17:  # 建議停損 Q
                    c.fill = mkfill(C_STOPBG)
                    c.font = mkfont('Calibri', 10, True, C_ACCENT)
                    c.number_format = '#,##0.00'
                elif ci == 20:  # 觸發訊號 T
                    c.alignment = ALN_LCW
                    c.font      = mkfont('Calibri', 10, True, '6E2F00')
                else:
                    c.font = mkfont('Calibri', 10, bold=(ci == 2), color='2C3E50')

        if RESULTS:
            last_d = 3 + len(RESULTS)
            tbl2 = Table(displayName='DetailTable', ref=f'B3:T{last_d}')
            tbl2.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
            ws2.add_table(tbl2)
            ws2.conditional_formatting.add(
                f'K4:K{last_d}',   # 布林頻寬移至 K 欄
                ColorScaleRule(
                    start_type='min',       start_color='63BE7B',
                    mid_type='percentile',  mid_value=50, mid_color='FFEB84',
                    end_type='max',         end_color='F8696B'
                ))
            ws2.conditional_formatting.add(
                f'N4:N{last_d}',   # 量比移至 N 欄
                DataBarRule(start_type='min', end_type='max', color='638EC6'))

        # 產業別欄(D) 依實際內容自動調整最佳欄寬（與 Sheet1 同步）
        max_ind_len_ws2 = max(
            (len(str(res['industry'])) for res in RESULTS),
            default=4
        )
        ws2.column_dimensions['D'].width = max(10, min(max_ind_len_ws2 * 2 + 4, 30))

        ws2.freeze_panes = 'E4'   # 凍結至產業別右側

        # ──────────────────────────────────────────────
        # Sheet 3：個股 K 線圖
        # 版面計算：
        #   figsize=(16,9) @dpi=150 → 實際圖片 2400×1350px
        #   插入 Excel 顯示尺寸 800×450px
        #   450px / (15.5pt×1.333px/pt) ≈ 21.8列 → 取 22列
        #   空白分隔列高 = 30pt ← 本次修改重點
        # ──────────────────────────────────────────────
        ws3 = wb.create_sheet('個股線型圖')
        ws3.merge_cells('B2:O2')
        c = ws3['B2']
        c.value     = f'個股技術分析線型圖 V3.1 ｜ 日期：{today_str} ｜ 共 {len(RESULTS)} 檔'
        c.font      = mkfont('Calibri', 13, True, 'FFFFFF')
        c.fill      = mkfill(C_NAVY)
        c.alignment = ALN_CC
        ws3.row_dimensions[2].height = 28
        ws3.row_dimensions[3].height = 4

        for ci in range(2, 18):
            ws3.column_dimensions[get_column_letter(ci)].width = 13

        # ── 版面參數 ──
        IMG_DISPLAY_W = 800   # px
        IMG_DISPLAY_H = 450   # px
        IMG_ROWS      = 22    # 圖片佔列數
        IMG_ROW_HT    = 15.5  # pt
        GAP_ROW_HT    = 30    # ★ 圖表間空白列高度（本次修正為 30pt）
        UNIT          = 1 + IMG_ROWS + 1  # 每組 = 1標題 + 22圖 + 1空白 = 24列

        chart_start = 4

        for idx, res in enumerate(RESULTS if excel_charts else []):
            print(f'  繪製圖表 [{idx+1}/{len(RESULTS)}] {res["code"]} {res["name"]}')

            img_path = os.path.join(IMG_DIR, f'chart_{res["code"]}.png')
            draw_stock_chart(res, img_path)

            base    = chart_start + idx * UNIT
            lbl_row = base

            # 標題列
            ws3.row_dimensions[lbl_row].height = 20
            ws3.merge_cells(f'B{lbl_row}:O{lbl_row}')
            c_lbl = ws3.cell(
                lbl_row, 2,
                f"  {idx+1}. {res['code']} {res['name']}  "
                f"收盤：{res['close']:.2f}  停損：{res['stop_loss']:.2f}  "
                f"觸發：{res['trigger']}"
            )
            c_lbl.font      = mkfont('Calibri', 10, True, 'FFFFFF')
            c_lbl.fill      = mkfill(C_TEAL)
            c_lbl.alignment = ALN_LC

            # 圖片列
            for rr in range(lbl_row + 1, lbl_row + IMG_ROWS + 1):
                ws3.row_dimensions[rr].height = IMG_ROW_HT

            # ★ 空白分隔列：30pt
            ws3.row_dimensions[lbl_row + IMG_ROWS + 1].height = GAP_ROW_HT

            # 插入圖片
            img_xl        = XLImage(img_path)
            img_xl.width  = IMG_DISPLAY_W
            img_xl.height = IMG_DISPLAY_H
            ws3.add_image(img_xl, f'B{lbl_row + 1}')

        ws3.freeze_panes = 'B3'

        # ===============================================================
        # 7. 儲存 Excel 並產生互動式線圖
        # ===============================================================
        wb.save(OUTPUT_FILE)
        size_kb = os.path.getsize(OUTPUT_FILE) / 1024
        print(f'\n✅ Excel 報告已儲存：{OUTPUT_FILE}（約 {size_kb:.0f} KB）')

    # ── 互動式 HTML 線圖 ─────────────────────────────────────────
    print('\n🎨 [3/3] 產生互動式個股技術線圖（HTML）...')
    html_path = build_interactive_html(
        RESULTS, today_str, output_dir, now,
        link_base=link_base, plotly_cdn=plotly_cdn,
        chart_years=chart_years, excel_url=excel_url,
    )

    # 排程要的是一個固定的檔名（`index.html`），因為下游——tw-six-metrics 的建站
    # 流程——是照名字來拿的；而帶時間戳的那一份是給人存檔用的。兩份同樣的內容，
    # 兩種不同的用途，複製一次比讓其中一方去猜另一方的檔名可靠。
    index_path = ''
    if html_path and index_copy:
        index_path = index_copy if os.path.isabs(index_copy) else os.path.join(
            os.getcwd(), index_copy)
        os.makedirs(os.path.dirname(os.path.abspath(index_path)) or '.', exist_ok=True)
        with open(html_path, encoding='utf-8') as src, \
             open(index_path, 'w', encoding='utf-8') as dst:
            dst.write(src.read())
        print(f'📄 已複製一份為：{index_path}')

    # 本機雙擊執行時把產出打開；排程上 `open_when_done` 是 False——runner 沒有
    # 桌面，`xdg-open` 只會噴一行錯誤，而且沒有人會按那個 Enter。
    if open_when_done:
        for target in (OUTPUT_FILE, html_path):
            if not target:
                continue
            try:
                if platform.system() == 'Windows':
                    os.startfile(target)                      # noqa: S606
                elif platform.system() == 'Darwin':
                    _sp.Popen(['open', target])
                else:
                    _sp.Popen(['xdg-open', target])
                print(f'📂 已開啟：{target}')
            except Exception as e:
                print(f'⚠️ 自動開啟失敗（{target}）：{e}')

    print()
    print('=' * 60)
    print(f'  篩選完成！通過 {len(RESULTS)} 檔，產出在 {output_dir}')
    print('=' * 60)
    return {
        'results': RESULTS,
        'count': len(RESULTS),
        'scanned': len(TICKERS),
        'date': today_str,
        'xlsx': OUTPUT_FILE,
        'html': html_path or '',
        'index': index_path,
    }





# ===============================================================
# 互動式個股技術線圖 v4 ─ script[type=application/json] 懶載入
# ===============================================================
def build_interactive_html(results, today_str, output_dir, now=None, *,
                           link_base='', plotly_cdn=True, chart_years=2.0,
                           excel_url=''):
    """產生互動線圖那一份 HTML，回傳檔案路徑。

    和本機版的三個差別，全都是因為這一份要放上網、給手機開：

    * ``plotly_cdn=True``：plotly.js 走 CDN 而不是內嵌。內嵌那 3.5 MB 讓每一位
      讀者每天重新下載一次同一份函式庫，而 CDN 上那一份會被瀏覽器快取；代價是
      離線打不開，所以本機版留著 ``plotly_cdn=False``。
    * ``chart_years``：K 線只留最近這幾年。指標全都是在**完整**歷史上算完才裁的，
      所以線的形狀和數值不會變，變的只有你能往左捲多遠。
    * ``link_base``：每一檔多一個連往〔六大財務指標評等〕個股頁的連結。技術面
      挑出來的標的，下一個問題一定是「這家公司體質怎麼樣」，而那個答案就在
      隔壁那個網站上。
    """
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
        import plotly.io as pio
        import json as _json, re as _re, math as _math
    except ImportError:
        print('plotly 未安裝，跳過互動線圖'); return None
    if not results: return None

    BG='#0d1117'; AXES='#161b22'; GRID='#30363d'; TEXT='#c9d1d9'
    MA20C='#f0c27f'; MA60C='#ff7b72'; BBC='#58a6ff'; STOPC='#ff4500'
    VUP='#3fb950'; VDN='#f85149'; VMAC='#ffa657'

    def sf(v, d=2):
        try:
            f=float(v); return 0.0 if (_math.isnan(f) or _math.isinf(f)) else round(f,d)
        except: return 0.0

    # ── 取 plotly.js 本體 ─────────────────────────────────────────
    # CDN 版把整份報告從 12 MB 壓到 1 MB 出頭。子資源完整性（SRI）雜湊寫死，
    # 這樣就算 CDN 上那個檔案被換掉，瀏覽器也會拒絕執行而不是照跑。
    if plotly_cdn:
        pljs = ('<script src="https://cdn.plot.ly/plotly-2.35.2.min.js" '
                'charset="utf-8"></script>')
    else:
        try:
            import plotly.offline as pyo
            plotly_js_src = pyo.get_plotlyjs()
            pljs = '<script>' + plotly_js_src + '</script>'
        except Exception as e:
            print(f'⚠️ plotly.js 內嵌失敗，改用 CDN：{e}')
            pljs = '<script src="https://cdn.plot.ly/plotly-2.35.2.min.js"></script>'

    #: 一年抓 252 個交易日。裁太少會讓「2年」那顆按鈕按下去看到一片空白，
    #: 所以多留半年的緩衝。
    keep_rows = int(chart_years * 252) + 130 if chart_years else 0

    stock_infos  = []
    fig_json_tags = []

    for idx, res in enumerate(results):
        df    = res['_df'].copy()          # 由 rangeselector 控制顯示範圍
        # 指標是在完整歷史上算完的（見 screen_stock），這裡才裁——先裁再算會讓
        # 最前面 60 根 K 棒的 60MA 變成 NaN，圖上就是一截斷掉的線。
        if keep_rows and len(df) > keep_rows:
            df = df.iloc[-keep_rows:]
        ma20  = res['_ma20'].reindex(df.index).ffill().bfill()
        ma60  = res['_ma60'].reindex(df.index).ffill().bfill()
        bup   = res['_boll_up'].reindex(df.index).ffill().bfill()
        bmid  = res['_boll_mid'].reindex(df.index).ffill().bfill()
        bdn   = res['_boll_dn'].reindex(df.index).ffill().bfill()
        vol   = (df['Volume'] / 1000).round(0)
        vmean = vol.rolling(20).mean().ffill().bfill()
        vratio= (vol / vmean).round(2)
        dates = df.index.strftime('%Y-%m-%d').tolist()
        stop  = res['stop_loss']
        N     = len(dates)

        o_l=[sf(v) for v in df['Open']];   h_l=[sf(v) for v in df['High']]
        l_l=[sf(v) for v in df['Low']];    c_l=[sf(v) for v in df['Close']]
        m20=[sf(v) for v in ma20];         m60=[sf(v) for v in ma60]
        bu =[sf(v) for v in bup];          bm =[sf(v) for v in bmid]
        bd =[sf(v) for v in bdn];          vl =[sf(v,0) for v in vol]
        vm =[sf(v,0) for v in vmean];      vr =[sf(v) for v in vratio]

        # customdata：[日期, 開, 高, 低, 收, 20MA, 60MA, BB上, BB中, BB下, 量, 均量, 量比]
        cd = [[dates[i],o_l[i],h_l[i],l_l[i],c_l[i],
               m20[i],m60[i],bu[i],bm[i],bd[i],
               vl[i],vm[i],vr[i]] for i in range(N)]
        cd_v = [[dates[i],vl[i],vm[i],vr[i]] for i in range(N)]

        htpl = (
            '<b style="font-size:13px">%{customdata[0]}</b><br>'
            '<span style="color:#8b949e">─────────────────────</span><br>'
            '開 <b>%{customdata[1]:.2f}</b> '
            '高 <b style="color:#3fb950">%{customdata[2]:.2f}</b> '
            '低 <b style="color:#f85149">%{customdata[3]:.2f}</b> '
            '收 <b style="color:#58a6ff">%{customdata[4]:.2f}</b><br>'
            '<span style="color:#f0c27f">20MA&nbsp;%{customdata[5]:.2f}</span>&nbsp;'
            '<span style="color:#ff7b72">60MA&nbsp;%{customdata[6]:.2f}</span><br>'
            '布林上&nbsp;%{customdata[7]:.2f}&nbsp;'
            '中&nbsp;%{customdata[8]:.2f}&nbsp;'
            '下&nbsp;%{customdata[9]:.2f}<br>'
            '<span style="color:#8b949e">─────────────────────</span><br>'
            '<span style="color:#3fb950">成交量 %{customdata[10]:,.0f} 張</span>  '
            '均量 %{customdata[11]:,.0f} 張  '
            '<span style="color:#ffa657">量比&nbsp;%{customdata[12]:.2f}x</span>'
            '<extra></extra>'
        )

        fig = make_subplots(rows=2, cols=1, shared_xaxes=True,
                            row_heights=[0.68, 0.32], vertical_spacing=0.03)

        # ① BB 上軌（先加，作為 tonexty 的填色目標）
        fig.add_trace(go.Scatter(
            x=dates, y=bu, mode='lines', name='布林上軌',
            line=dict(color=BBC, width=1.2, dash='dot'),
            showlegend=True, hoverinfo='skip',
        ), row=1, col=1)

        # ② BB 下軌 + tonexty 填色（填向上一條同軸 trace）
        fig.add_trace(go.Scatter(
            x=dates, y=bd, mode='lines', name='布林下軌',
            line=dict(color=BBC, width=1.2, dash='dot'),
            fill='tonexty', fillcolor='rgba(88,166,255,0.08)',
            showlegend=False, hoverinfo='skip',
        ), row=1, col=1)

        # ③ BB 中軌
        fig.add_trace(go.Scatter(x=dates, y=bm, mode='lines', name='布林中軌',
            line=dict(color='#8b949e', width=0.8, dash='dot'),
            showlegend=False, hoverinfo='skip'), row=1, col=1)

        # ④ K 棒
        fig.add_trace(go.Candlestick(
            x=dates, open=o_l, high=h_l, low=l_l, close=c_l,
            increasing_line_color=VUP, decreasing_line_color=VDN,
            increasing_fillcolor=VUP,  decreasing_fillcolor=VDN,
            name='K棒', showlegend=True,
            hoverinfo='skip', line=dict(width=1),
        ), row=1, col=1)

        # ⑤ 20MA
        fig.add_trace(go.Scatter(x=dates, y=m20, mode='lines', name='20MA 月線',
            line=dict(color=MA20C, width=2.0),
            showlegend=True, hoverinfo='skip'), row=1, col=1)

        # ⑥ 60MA
        fig.add_trace(go.Scatter(x=dates, y=m60, mode='lines', name='60MA 季線',
            line=dict(color=MA60C, width=2.0, dash='dash'),
            showlegend=True, hoverinfo='skip'), row=1, col=1)

        # ⑦ 停損線
        fig.add_trace(go.Scatter(
            x=[dates[0], dates[-1]], y=[stop, stop],
            mode='lines', name='建議停損 {:.2f}'.format(stop),
            line=dict(color=STOPC, width=1.5, dash='dashdot'),
            showlegend=True, hoverinfo='skip'), row=1, col=1)

        # ⑧ 成交量 Bar
        vcols=[VUP if c_l[i]>=o_l[i] else VDN for i in range(N)]
        fig.add_trace(go.Bar(x=dates, y=vl, name='成交量(張)',
            marker_color=vcols, opacity=0.85, showlegend=True,
            hoverinfo='skip'), row=2, col=1)

        # ⑨ 均量線
        fig.add_trace(go.Scatter(x=dates, y=vm, mode='lines', name='20日均量',
            line=dict(color=VMAC, width=1.5),
            showlegend=True, hoverinfo='skip'), row=2, col=1)

        # ⑩ 價格 invisible scatter（承載 hover tooltip）
        fig.add_trace(go.Scatter(
            x=dates, y=c_l, mode='markers',
            marker=dict(opacity=0, size=14, color='rgba(0,0,0,0)'),
            customdata=cd, hovertemplate=htpl,
            showlegend=False, name='__hover_price__',
        ), row=1, col=1)

        # ⑪ 量圖 invisible scatter（讓量圖也有 spike + tooltip 貫穿）
        htpl_v = ('成交量 <b style="color:#3fb950">%{customdata[1]:,.0f}</b> 張&nbsp;&nbsp;'
                  '均量 %{customdata[2]:,.0f} 張&nbsp;&nbsp;'
                  '<span style="color:#ffa657">量比 %{customdata[3]:.2f}x</span>'
                  '<extra></extra>')
        fig.add_trace(go.Scatter(
            x=dates, y=vl, mode='markers',
            marker=dict(opacity=0, size=14, color='rgba(0,0,0,0)'),
            customdata=cd_v, hovertemplate=htpl_v,
            showlegend=False, name='__hover_vol__',
        ), row=2, col=1)

        # ── Layout（用 update_xaxes 避免覆蓋 make_subplots 內部的 matches 配置）──
        # 注意：不在 update_xaxes 設 type='date'，避免與 xaxis2.matches='x' 衝突
        # plotly 會從 x 資料（YYYY-MM-DD 字串）自動偵測為 date 軸
        # 註：垂直十字線改由前端 JS（plotly_hover / plotly_unhover）以 shape 手動繪製，
        # 使其能真正貫穿上（K棒）下（成交量）兩張子圖；此處停用 plotly 原生 x 軸 spike，
        # 避免原生 spike（僅侷限於單一子圖繪圖區）與自繪十字線重疊顯示。
        fig.update_xaxes(
            showspikes=False,
            gridcolor=GRID, gridwidth=0.5, tickfont=dict(color=TEXT),
            showgrid=True,
        )
        fig.update_yaxes(
            showspikes=True, spikecolor='#8b949e', spikethickness=1,
            spikedash='dot', gridcolor=GRID, gridwidth=0.5,
            tickfont=dict(color=TEXT), zeroline=False,
        )
        # rangeslider 也用 update_xaxes 停用，避免 update_layout shorthand 的副作用
        fig.update_xaxes(rangeslider_visible=False)
        fig.update_layout(
            paper_bgcolor=BG, plot_bgcolor=AXES,
            font=dict(family='Microsoft JhengHei, Arial', color=TEXT, size=12),
            hovermode='x unified',
            hoverlabel=dict(bgcolor='rgba(13,17,23,0.95)', bordercolor='#444',
                            font=dict(color=TEXT, size=12,
                                      family='Microsoft JhengHei, Arial'),
                            align='left', namelength=0),
            legend=dict(orientation='h', x=0.5, xanchor='center', y=1.02,
                        bgcolor='rgba(22,27,34,0.9)', bordercolor=GRID,
                        borderwidth=1, font=dict(size=11)),
            margin=dict(t=46, b=26, l=64, r=26),
            height=760, dragmode='pan',
            autosize=True,
        )
        fig.update_yaxes(title_text='價格 (元)',  row=1, col=1,
                         title_font=dict(color='#8b949e', size=11))
        fig.update_yaxes(title_text='成交量(張)', row=2, col=1,
                         title_font=dict(color='#8b949e', size=11))

        # 將 figure JSON 存成 <script type="application/json"> tag
        fig_json = pio.to_json(fig)
        # 防止 </script> 提前關閉：把 </ 轉義
        fig_json_safe = fig_json.replace('</', '<\\/')
        fig_json_tags.append(
            '<script type="application/json" id="fig-{i}">{j}</script>'.format(
                i=idx, j=fig_json_safe)
        )

        # 漲跌與漲跌幅：以最近兩個交易日收盤價計算，供左側清單顯示
        if N >= 2 and c_l[-2]:
            chg     = round(c_l[-1] - c_l[-2], 2)
            chg_pct = round(chg / c_l[-2] * 100, 2)
        else:
            chg, chg_pct = 0.0, 0.0

        stock_infos.append({
            'code': res['code'],    'name': res['name'],
            'ind':  res.get('industry',''),
            'close':res['close'],   'ma20': res['ma20_last'],
            'ma60': res['ma60_last'],'stop': res['stop_loss'],
            'vratio':res['vol_ratio'],'bw': res['boll_bw_pct'],
            'atr':  res['atr14'],   'trigger': res['trigger'],
            'chg':  chg,            'chg_pct': chg_pct,
            # 連往〔六大財務指標評等〕的個股頁。`link_base` 空的時候是空字串，
            # 前端就不畫那顆按鈕——一個指向 undefined.html 的連結比沒有連結糟。
            'url':  (link_base.rstrip('/') + '/' + res['code'] + '.html')
                    if link_base else '',
        })

    # ── 組裝 HTML ────────────────────────────────────────────────
    n      = len(results)
    # 檔名時間戳記：與 Excel 相同規則 YYYYMMDD_HHMM，同天多次執行不互蓋
    if now is None:
        now = datetime.datetime.now()
    ts     = now.strftime('%Y_%m_%d_%H%M')   # e.g. 2026_08_06_1430
    ts_display = today_str.replace('-', '/')  # 用於 HTML 顯示
    info_js = _json.dumps(stock_infos, ensure_ascii=False)

    def _esc(t):
        return (str(t).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'))

    # 側欄卡片右上角那個 ↗ 直接開個股頁。它在 `onclick="showChart(i)"` 的卡片
    # **裡面**，所以要擋掉事件冒泡——不然點連結會順便把圖表也切過去，讀者回來
    # 之後看到的是另一檔。
    def _sb_link(s):
        if not s.get('url'):
            return ''
        return ('<a class="nb-ext" href="{u}" target="_blank" rel="noopener" '
                'title="看這一檔的六大財務指標評等" '
                'onclick="event.stopPropagation()">六大&#x2197;</a>').format(u=s['url'])

    nav_btns = '\n'.join(
        '<div class="nb" id="btn-{i}" onclick="showChart({i})">'
        '<div class="nb-top">'
        '<span class="nb-code">{code}</span>'
        '<span class="nb-close">{close:.2f}</span>'
        '</div>'
        '<div class="nb-bot">'
        '<span class="nb-name">{name}</span>'
        '<span class="nb-pct {pcls}">{sign}{pct:.2f}%</span>'
        '</div>'
        '<div class="nb-tagrow"><span class="nb-tag">{trig}</span>{ext}</div>'
        '</div>'.format(
            i=i, code=s['code'], name=_esc(s['name']),
            close=s['close'], pct=abs(s['chg_pct']),
            sign='▲' if s['chg_pct'] > 0 else ('▼' if s['chg_pct'] < 0 else ''),
            pcls='up' if s['chg_pct'] > 0 else ('dn' if s['chg_pct'] < 0 else 'fl'),
            ext=_sb_link(s),
            trig=_esc(s['trigger'])[:22] + ('…' if len(s['trigger']) > 22 else ''))
        for i, s in enumerate(stock_infos)
    )
    chart_divs = '\n'.join(
        '<div class="cw" id="cw-{i}">'
        '<div class="badge-row" id="badges-{i}"></div>'
        '<div class="trig-line" id="trig-{i}"></div>'
        '<div class="plot" id="plot-{i}"><p class="ld">&#9203; 載入中...</p></div>'
        '<div id="rb-{i}">'
        '<span>時間範圍：</span>'
        '<button class="rbtn"     onclick="setRange(1,0,this,{i})">1月</button>'
        '<button class="rbtn rba" onclick="setRange(3,0,this,{i})">3月</button>'
        '<button class="rbtn"     onclick="setRange(6,0,this,{i})">6月</button>'
        '<button class="rbtn"     onclick="setRange(0,1,this,{i})">1年</button>'
        '<button class="rbtn"     onclick="setRange(0,{y},this,{i})">{y}年</button>'
        '</div>'
        '</div>'.format(i=i, y=int(chart_years) if chart_years else 2)
        for i in range(n)
    )

    css = (
        '*{box-sizing:border-box;margin:0;padding:0}'
        'body{background:#0d1117;color:#c9d1d9;'
            'font-family:"Microsoft JhengHei",Arial,sans-serif;overflow-x:hidden}'
        # ── 頂部標題列（桌機顯示，手機隱藏）───────────────────
        '#hd{background:#161b22;border-bottom:1px solid #30363d;'
            'padding:9px 16px;display:flex;align-items:center;gap:14px;flex-wrap:wrap}'
        '#hd h1{font-size:15px;color:#e6edf3;white-space:nowrap}'
        '.meta{font-size:12px;color:#8b949e}'
        '.tip{font-size:12px;color:#f0c27f;margin-left:auto}'
        # ── 主體：左側清單 + 右側圖表，左右分欄、各自獨立捲動 ───
        'html,body{height:100%;overflow:hidden}'
        ':root{--hd-h:46px}'
        '#topbar{position:sticky;top:0;z-index:30;background:#0d1117;'
            'box-shadow:0 2px 8px rgba(0,0,0,0.35)}'
        '#main{display:flex;align-items:stretch;'
            'height:calc(100vh - var(--hd-h, 46px))}'
        '#sidebar{width:300px;flex:0 0 300px;background:#0d1117;'
            'border-right:1px solid #30363d;height:100%;overflow-y:auto}'
        '#sb-list{padding:6px}'
        '#chartcol{flex:1 1 auto;min-width:0;height:100%;'
            'display:flex;flex-direction:column;overflow-y:auto}'
        # ── 左側清單卡片 ────────────────────────────────────────
        '.nb{background:#161b22;border:1px solid #30363d;'
            'color:#c9d1d9;cursor:pointer;padding:10px 12px;margin:5px 4px;'
            'border-radius:8px;font-family:inherit;'
            'transition:background .15s,border-color .15s}'
        '.nb:hover{background:#1c2530;border-color:#388bfd}'
        '.nb.act{background:#132a45;border-color:#58a6ff}'
        '.nb-top{display:flex;justify-content:space-between;align-items:baseline}'
        '.nb-code{font-size:16px;font-weight:bold;color:#e6edf3}'
        '.nb-close{font-size:16px;font-weight:bold;color:#e6edf3}'
        '.nb-bot{display:flex;justify-content:space-between;align-items:baseline;'
            'margin-top:2px;font-size:12px;color:#8b949e}'
        '.nb-pct{font-weight:bold}'
        '.nb-pct.up{color:#3fb950}'
        '.nb-pct.dn{color:#f85149}'
        '.nb-pct.fl{color:#8b949e}'
        '.nb-tagrow{display:flex;align-items:center;gap:6px;margin-top:6px}'
        '.nb-tag{display:inline-block;padding:2px 8px;'
            'border-radius:11px;font-size:10.5px;background:rgba(240,194,127,.14);'
            'color:#f0c27f;border:1px solid rgba(240,194,127,.32);'
            'min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}'
        # 側欄卡片上那個「六大↗」。刻意做得比觸發訊號那顆淡一點：它是一條出口，
        # 不是這一頁在講的事。
        '.nb-ext{margin-left:auto;flex:0 0 auto;padding:2px 8px;border-radius:11px;'
            'font-size:10.5px;text-decoration:none;background:rgba(88,166,255,.12);'
            'color:#58a6ff;border:1px solid rgba(88,166,255,.30);white-space:nowrap}'
        '.nb-ext:hover{background:rgba(88,166,255,.24);border-color:#58a6ff}'
        # 圖表區那顆同樣的出口，做成一顆徽章，跟 20MA／停損那幾顆並排。
        'a.badge.link{text-decoration:none;background:rgba(88,166,255,.16);'
            'color:#79c0ff;border-color:rgba(88,166,255,.42)}'
        'a.badge.link:hover{background:rgba(88,166,255,.30);border-color:#58a6ff}'
        # 頂端那條 Excel 下載連結。
        '#hd a.dl{color:#7ee787;text-decoration:none;font-size:12px;'
            'border:1px solid rgba(126,231,135,.32);border-radius:12px;'
            'padding:3px 10px;white-space:nowrap}'
        '#hd a.dl:hover{background:rgba(126,231,135,.14)}'
        # ── 圖表區內：徽章列 + 觸發訊號 + 圖表本體 + 時間範圍列 ─
        '.cw{display:none;width:100%;height:100%;padding:10px 16px 12px;'
            'flex-direction:column;min-height:0}'
        '.cw.act{display:flex}'
        '.badge-row{display:flex;flex-wrap:wrap;gap:8px;align-items:center;'
            'margin-bottom:6px;flex:0 0 auto}'
        '.badge{display:inline-flex;align-items:center;padding:5px 13px;'
            'border-radius:16px;font-size:12.5px;font-weight:600;'
            'background:rgba(88,166,255,.12);color:#58a6ff;'
            'border:1px solid rgba(88,166,255,.3);white-space:nowrap}'
        '.badge.ind{background:rgba(88,166,255,.16);color:#79c0ff}'
        '.badge.stop{background:rgba(248,81,73,.14);color:#f85149;'
            'border-color:rgba(248,81,73,.35)}'
        '.badge.gold{background:rgba(240,194,127,.14);color:#f0c27f;'
            'border-color:rgba(240,194,127,.35)}'
        '.trig-line{font-size:13px;color:#f0c27f;margin:2px 0 8px;'
            'display:flex;align-items:center;gap:6px;flex:0 0 auto}'
        '.trig-arrow{color:#ffa657}'
        '.ld{text-align:center;padding:80px;color:#8b949e;font-size:15px}'
        '.plot{width:100%;flex:1 1 auto;min-height:0}'
        '[id^="rb-"]{display:flex;align-items:center;gap:6px;flex-wrap:wrap;'
            'justify-content:center;margin-top:8px;flex:0 0 auto}'
        '[id^="rb-"] span{font-size:11px;color:#8b949e;margin-right:4px}'
        '.rbtn{background:#21262d;border:1px solid #30363d;color:#c9d1d9;'
            'cursor:pointer;padding:4px 13px;border-radius:16px;'
            'font-family:inherit;font-size:11px;transition:background .12s}'
        '.rbtn:hover{background:#30363d;border-color:#58a6ff}'
        '.rbtn.rba{background:#1f6feb;border-color:#388bfd;color:#fff}'
        # ── RWD：手機隱藏頂部標題列，改為單欄堆疊（保留頁面捲動）─
        '@media(max-width:760px){'
            'html,body{overflow:auto}'
            '#hd{display:none}'
            '#main{flex-direction:column;height:auto}'
            '#sidebar{width:100%;flex:0 0 auto;max-height:220px;height:auto;'
                'border-right:none;border-bottom:1px solid #30363d}'
            '#chartcol{height:auto;overflow-y:visible}'
            '.cw.act{height:auto}'
            '.plot{height:520px}'
        '}'
    )

    js = r"""
const INFO = INFO_JS_PLACEHOLDER;
const rendered = {};
const PLY_CFG = {
  displayModeBar: true, displaylogo: false, scrollZoom: true,
  modeBarButtonsToRemove: ['lasso2d','select2d'],
  toImageButtonOptions: {
    format:'png', filename:'twstock_TS_PH', height:760, width:1400, scale:2
  }
};

/* ── 依實際 #hd 高度設定 CSS 變數，讓 #main 精準填滿剩餘視窗高度 ── */
function syncHdHeight() {
  var hd = document.getElementById('topbar');
  var h = hd ? hd.offsetHeight : 0;
  document.documentElement.style.setProperty('--hd-h', h + 'px');
}

/* ── 讓圖表填滿 .plot 容器的實際像素高度（而非寫死的 760）──── */
function fitPlotSize(div) {
  if (!div) return;
  var w = div.clientWidth, h = div.clientHeight;
  if (w > 0 && h > 0) Plotly.relayout(div, {width: w, height: h});
}

/* ── 台灣時間（UTC+8）─────────────────────────────────────── */
function twNow() {
  var now = new Date();
  /* 加上 UTC+8 偏移，再修正本機時區偏移，得到台灣當地時間 */
  return new Date(now.getTime() + (8 * 60 + now.getTimezoneOffset()) * 60000);
}
function toISO(d) { return d.toISOString().slice(0, 10); }

/* ── 設定時間範圍（台灣時間往回推）──────────────────────── */
function setRange(months, years, btn, idx) {
  var rbWrap = document.getElementById('rb-' + idx);
  if (rbWrap) rbWrap.querySelectorAll('.rbtn').forEach(function(b){ b.classList.remove('rba'); });
  if (btn) btn.classList.add('rba');

  var wrap = document.getElementById('plot-' + idx);
  if (!wrap || !rendered[idx]) return;

  var end   = twNow();
  var start = new Date(end);
  if (months) start.setMonth(start.getMonth() - months);
  if (years)  start.setFullYear(start.getFullYear() - years);

  /* 只設定 X 軸範圍；Y 軸交由 attachAutoY 的 relayout 監聽器，
     依「目前可視範圍內的實際最高/最低值」自動計算最佳上下界，避免留白過多 */
  Plotly.relayout(wrap, {
    /* 主圖 xaxis.matches='x2'，故 range 必須設在 xaxis2 才會真正生效並同步 */
    'xaxis2.range[0]': toISO(start),
    'xaxis2.range[1]': toISO(end)
  });
}

/* ── 解析 figure JSON（安全，不受 HTML 標籤干擾）─────────── */
function getChartFig(idx) {
  var el = document.getElementById('fig-' + idx);
  if (!el) { console.error('找不到 fig-' + idx); return null; }
  try { return JSON.parse(el.textContent); }
  catch(e) { console.error('JSON 解析失敗 fig-' + idx, e); return null; }
}

/* ── 資訊欄更新 ───────────────────────────────────────────── */
function updateIB(i) {
  var s = INFO[i];
  var badges = document.getElementById('badges-' + i);
  if (badges && !badges.dataset.filled) {
    badges.innerHTML =
      '<span class="badge ind">' + (s.ind||'\u2014') + '</span>' +
      '<span class="badge gold">20MA ' + s.ma20.toFixed(2) + '</span>' +
      '<span class="badge gold">60MA ' + s.ma60.toFixed(2) + '</span>' +
      '<span class="badge stop">\u505c\u640d ' + s.stop.toFixed(2) + '</span>' +
      '<span class="badge">\u91cf\u6bd4 ' + s.vratio.toFixed(2) + 'x</span>' +
      '<span class="badge">\u5e03\u6797\u5bec ' + s.bw.toFixed(1) + '%</span>' +
      '<span class="badge gold">ATR ' + s.atr.toFixed(2) + '</span>' +
      /* \u6280\u8853\u9762\u770b\u5b8c\uff0c\u4e0b\u4e00\u500b\u554f\u984c\u662f\u9019\u5bb6\u516c\u53f8\u7684\u9ad4\u8cea\u2014\u2014\u90a3\u500b\u7b54\u6848\u5728\u9694\u58c1\u90a3\u500b\u7db2\u7ad9\u4e0a\u3002
         target="_blank"\uff1a\u9019\u4e00\u9801\u7684\u5716\u8868\u662f\u6709\u72c0\u614b\u7684\uff08\u9078\u4e86\u54ea\u4e00\u6a94\u3001\u62c9\u5230\u54ea\u500b\u7bc4\u570d\uff09\uff0c
         \u5728\u539f\u5730\u8df3\u8d70\u6703\u628a\u90a3\u4e9b\u5168\u90e8\u4e1f\u6389\u3002 */
      (s.url
        ? '<a class="badge link" href="' + s.url + '" target="_blank" ' +
          'rel="noopener">\u516d\u5927\u8ca1\u52d9\u6307\u6a19\u8a55\u7b49 ' +
          s.code + '\u2197</a>'
        : '');
    var trig = document.getElementById('trig-' + i);
    if (trig) trig.innerHTML =
      '<span class="trig-arrow">\u25b6\ufe0e</span>\u00a0' + s.trigger;
    badges.dataset.filled = '1';
  }
}

/* ── 貫穿上下兩圖的十字線（自繪 shape，涵蓋 paper 全高）───── */
var xlineLast = null;
function drawXLine(div, xv) {
  if (xlineLast === xv) return;   /* 同一天不重複觸發 relayout */
  xlineLast = xv;
  Plotly.relayout(div, {
    shapes: [{
      type: 'line', xref: 'x2', yref: 'paper',
      x0: xv, x1: xv, y0: 0, y1: 1,
      line: { color: '#58a6ff', width: 1, dash: 'solid' },
      layer: 'above'
    }]
  });
}
function clearXLine(div) {
  xlineLast = null;
  Plotly.relayout(div, { shapes: [] });
}

/* ── Hover：更新可複製資料列 + 繪製貫穿十字線 ─────────────── */
function attachHover(div) {
  div.on('plotly_hover', function(data) {
    var pts = data.points || [];
    var pp = null;
    for (var k = 0; k < pts.length; k++) {
      if (pts[k].customdata && pts[k].customdata.length >= 10) { pp = pts[k]; break; }
    }
    if (!pp) return;
    var cd = pp.customdata;
    drawXLine(div, cd[0]);
  });
  div.on('plotly_unhover', function() { clearXLine(div); });
}

/* ── 依目前可視範圍內的實際數值，計算最佳 Y 軸上下界 ─────── */
function padRange(lo, hi, ratio) {
  if (lo === hi) {
    var eps = (Math.abs(lo) || 1) * 0.05;
    return [lo - eps, hi + eps];
  }
  var pad = (hi - lo) * ratio;
  return [lo - pad, hi + pad];
}
function computeVisibleYRange(div, xStart, xEnd) {
  var loP = Infinity, hiP = -Infinity, loV = Infinity, hiV = -Infinity;
  (div.data || []).forEach(function(tr) {
    if (tr.name === '__hover_price__' || tr.name === '__hover_vol__') return;
    var xs = tr.x || [];
    var isVol = (tr.yaxis === 'y2');
    for (var i = 0; i < xs.length; i++) {
      var xd = new Date(xs[i]);
      if (xd < xStart || xd > xEnd) continue;
      var vals;
      if (tr.type === 'candlestick') vals = [tr.high[i], tr.low[i]];
      else if (tr.y) vals = [tr.y[i]];
      else continue;
      for (var j = 0; j < vals.length; j++) {
        var v = vals[j];
        if (v === null || v === undefined || isNaN(v)) continue;
        if (isVol) { if (v < loV) loV = v; if (v > hiV) hiV = v; }
        else       { if (v < loP) loP = v; if (v > hiP) hiP = v; }
      }
    }
  });
  return {
    y1: (isFinite(loP) && isFinite(hiP)) ? padRange(loP, hiP, 0.06) : null,
    /* 量圖固定以 0 為底，只在頂端留一點空間，符合成交量圖慣例 */
    y2: (isFinite(loV) && isFinite(hiV)) ? [0, hiV * 1.12] : null
  };
}

/* ── 縮放/拖曳後自動調整 Y 軸範圍（依目前可視範圍內的資料自動縮放）───
   用 setTimeout 延後到下一個 tick 才真正套用，並用世代計數器捨棄過期
   的計算結果，避免在拖曳/縮放手勢「進行中」對圖表下達 relayout 而
   造成畫面彈回原位。 ────────────────────────────────────────── */
var autoYGen = 0;
function attachAutoY(div) {
  div.on('plotly_relayout', function(ev) {
    var r0 = ev['xaxis2.range[0]'] !== undefined ? ev['xaxis2.range[0]'] : ev['xaxis.range[0]'];
    var r1 = ev['xaxis2.range[1]'] !== undefined ? ev['xaxis2.range[1]'] : ev['xaxis.range[1]'];
    var xAutorange = (ev['xaxis2.autorange'] === true || ev['xaxis.autorange'] === true);
    if (r0 === undefined && r1 === undefined && !xAutorange) return;

    var xStart, xEnd;
    if (xAutorange && (r0 === undefined || r1 === undefined)) {
      /* 雙擊重置等情況：無明確範圍，視為顯示全部資料 */
      xStart = new Date(-8640000000000000);
      xEnd   = new Date(8640000000000000);
    } else {
      xStart = new Date(r0);
      xEnd   = new Date(r1);
    }

    var myGen = ++autoYGen;
    setTimeout(function() {
      if (myGen !== autoYGen) return;   /* 期間又有新的縮放/拖曳發生，捨棄這次結果 */
      var rng = computeVisibleYRange(div, xStart, xEnd);
      var upd = {};
      if (rng.y1) { upd['yaxis.range']  = rng.y1;  upd['yaxis.autorange']  = false; }
      if (rng.y2) { upd['yaxis2.range'] = rng.y2;  upd['yaxis2.autorange'] = false; }
      if (Object.keys(upd).length) Plotly.relayout(div, upd);
    }, 0);
  });
}

/* ── 切換個股 ─────────────────────────────────────────────── */
function showChart(idx) {
  document.querySelectorAll('.nb').forEach(function(b,i){ b.classList.toggle('act', i===idx); });
  document.querySelectorAll('.cw').forEach(function(el,i){ el.classList.toggle('act', i===idx); });
  updateIB(idx);

  var wrap = document.getElementById('plot-' + idx);
  if (!rendered[idx]) {
    rendered[idx] = true;
    var fig = getChartFig(idx);
    if (!fig) {
      wrap.innerHTML = '<p class="ld" style="color:#f85149">\u26a0 \u5716\u8868\u8cc7\u6599\u89e3\u6790\u5931\u6557</p>';
      return;
    }
    wrap.innerHTML = '';   /* 清空「載入中」佔位文字，避免殘留在圖表下方 */
    Plotly.newPlot(wrap, fig.data, fig.layout, PLY_CFG)
      .then(function() {
        fitPlotSize(wrap);
        attachHover(wrap);
        attachAutoY(wrap);
        /* 預設顯示 3 個月（台灣時間往回推），同時自動調整 Y 軸 */
        var rbWrap = document.getElementById('rb-' + idx);
        var activeBtn = rbWrap ? rbWrap.querySelector('.rbtn.rba') : null;
        if (activeBtn) activeBtn.click();
        else {
          /* 找第一個按鈕並觸發 1 月 */
          var first = rbWrap ? rbWrap.querySelector('.rbtn') : null;
          if (first) first.click();
        }
      })
      .catch(function(e) {
        wrap.innerHTML = '<p class="ld" style="color:#f85149">\u26a0 Plotly \u7573\u8b5c\u5931\u6557: ' + e.message + '</p>';
      });
  } else {
    fitPlotSize(wrap);
    /* 切換時重新套用目前選中的時間範圍 + 自動調整 Y */
    var rbWrap = document.getElementById('rb-' + idx);
    var activeBtn = rbWrap ? rbWrap.querySelector('.rbtn.rba') : null;
    if (activeBtn) activeBtn.click();
  }
}

window.addEventListener('resize', function() {
  syncHdHeight();
  var a = document.querySelector('.cw.act .plot');
  if (a) fitPlotSize(a);
});

document.addEventListener('DOMContentLoaded', function() { syncHdHeight(); showChart(0); });
"""
    js = js.replace('INFO_JS_PLACEHOLDER', info_js)\
           .replace('TS_PH', ts.replace('_', ''))

    parts = [
        '<!DOCTYPE html><html lang="zh-TW"><head>',
        '<meta charset="UTF-8">',
        '<meta name="viewport" content="width=device-width,initial-scale=1">',
        '<title>\u53f0\u80a1\u9806\u52e2\u4ea4\u6613\u4e92\u52d5\u7dda\u5716 ' + ts_display + '</title>',
        '<style>' + css + '</style>',
        pljs,
        '</head><body>',
        '<div id="topbar">',
        '<div id="hd">',
        '<h1>&#x1F1F9;&#x1F1FC; \u53f0\u80a1\u9806\u52e2\u4ea4\u6613\u7be9\u9078\u7cfb\u7d71 V3.1</h1>',
        '<div class="meta">\u7be9\u9078\u65e5\u671f\uff1a' + ts_display +
        '&nbsp;|&nbsp;\u5171&nbsp;<b style="color:#3fb950">' + str(n) +
        '</b>&nbsp;\u6a94\u901a\u904e</div>',
        '<div class="tip">&#x1F5B1;&#xFE0F; \u6ed1\u9f20\u79fb\u5165\u5716\u8868 \u2192 '
        '\u5373\u6642\u986f\u793a\u6240\u6709\u6307\u6a19\uff08\u542b\u91cf\u5716\uff09'
        '&nbsp;|&nbsp;\u5de6\u9375\u62d6\u66f3\u5e73\u79fb&nbsp;|&nbsp;\u6eda\u8f2a\u7e2e\u653e</div>',
        # Excel \u7248\uff08\u6bcf\u4e00\u6a94\u4e00\u5f35 K \u7dda\u5716\uff09\u5b58\u5728\u90a3\u4e00\u6b21\u6392\u7a0b\u7684 artifact \u88e1\u3002\u9023\u7d50\u6307\u5411\u90a3\u4e00\u6b21
        # \u7684\u57f7\u884c\u9801\u9762\uff0c\u800c\u4e0d\u662f\u4e00\u500b\u76f4\u63a5\u4e0b\u8f09\u7684\u7db2\u5740\u2014\u2014GitHub \u7684 artifact \u7db2\u5740\u662f\u7c3d\u904e\u7ae0\u3001
        # \u5e7e\u5206\u9418\u5c31\u904e\u671f\u7684\uff0c\u5beb\u6b7b\u5728\u5831\u544a\u88e1\u7b49\u65bc\u5beb\u6b7b\u4e00\u500b\u58de\u9023\u7d50\u3002
        #
        # artifact \u4fdd\u7559 30 \u5929\u3002\u904e\u4e86\u5c31\u4e0b\u8f09\u4e0d\u5230\uff0c\u9023\u7d50\u6703\u505c\u5728\u4e00\u500b\u300c\u5df2\u904e\u671f\u300d\u7684\u9801\u9762\u2014\u2014
        # \u9019\u4ef6\u4e8b\u5beb\u5728\u9023\u7d50\u65c1\u908a\uff0c\u8b80\u8005\u4e0d\u5fc5\u9ede\u4e0b\u53bb\u624d\u77e5\u9053\u3002
        ('<a class="dl" href="' + excel_url + '" target="_blank" rel="noopener" '
         'title="\u8a72\u6b21\u57f7\u884c\u7684 Artifacts \u5340\uff0c\u4fdd\u7559 30 \u5929">'
         '&#x2B07;&#xFE0F; Excel \u5831\u8868\uff0830 \u5929\u5167\uff09</a>')
        if excel_url else '',
        '</div>',
        '</div>',  # /#topbar
        '<div id="main">',
        '<div id="sidebar">',
        '<div id="sb-list">' + nav_btns + '</div>',
        '</div>',  # /#sidebar
        '<div id="chartcol">' + chart_divs + '</div>',
        '</div>',  # /#main
        '\n'.join(fig_json_tags),   # <-- 圖表資料放在獨立 script[type=application/json]
        '<script>' + js + '</script>',
        '</body></html>',
    ]

    html_path = os.path.join(output_dir, 'TW_Stock_Trend_Following_Trading_Filter_Result_V3.1_' + ts + '.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(''.join(parts))
    mb = os.path.getsize(html_path) / 1024 / 1024
    print('\u4e92\u52d5\u7dda\u5716\u5132\u5b58\uff1a{} ({:.1f} MB)'.format(html_path, mb))
    return html_path


if __name__ == '__main__':
    main()
